import os
from accounting_utils import (
    get_account_info_by_code,
    get_customer_posting_account_code,
    get_supplier_posting_account_code,
    get_account_id_by_code,
    create_account_transaction,
    create_double_entry,
    get_customer_account_id,
    get_supplier_account_id,
    create_customer_double_entry,
    create_supplier_double_entry,
    create_supplier_payment_entry,
    create_customer_collection_entry
)
from io import BytesIO
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from models import db, User, UserRole, Permission,SupplierPayment, ROLE_PERMISSIONS, Employee, Supplier, Product, PurchaseOrder, PurchaseItem, Customer, SaleOrder, SaleItem, Collection, CashTransaction, FreezeDeposit, Transaction, CashBox, JournalEntry, JournalDetail, DailyCashSummary, FinancialAccount, AccountTransaction
from datetime import datetime
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import json
from werkzeug.security import generate_password_hash, check_password_hash
import bcrypt  # إضافة هذا الاستيراد


app = Flask(__name__)

# إعدادات التطبيق الأساسية
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-change-this')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# إعداد قاعدة البيانات (دعم SQLite محلياً و PostgreSQL على Render)
database_url = os.environ.get('DATABASE_URL', 'sqlite:///thaljat_alsaleef.db')
if database_url and database_url.startswith('postgres://'):
    database_url = database_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = database_url

# تهيئة قاعدة البيانات
db.init_app(app)

# تهيئة نظام تسجيل الدخول
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

from datetime import datetime

@app.context_processor
def utility_processor():
    """جعل الدوال والمتغيرات متاحة في جميع القوالب"""

    def now(format='%Y-%m-%d %H:%M:%S'):
        return datetime.now().strftime(format)

    def format_number(value):
        """تنسيق الأرقام"""
        try:
            return f"{float(value):,.2f}"
        except:
            return value

    def format_currency(value):
        """تنسيق العملة"""
        try:
            return f"{float(value):,.2f} ر.ي"
        except:
            return value

    return {
        'datetime': datetime,
        'now': now,
        'current_date': datetime.now().strftime('%Y-%m-%d'),
        'current_time': datetime.now().strftime('%H:%M:%S'),
        'current_year': datetime.now().year,
        'format_number': format_number,
        'format_currency': format_currency
    }

def permission_required(permission):
    """ديكوراتور للتحقق من الصلاحيات"""

    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.has_permission(permission):
                flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)

        return decorated_function

    return decorator

import arabic_reshaper
from bidi.algorithm import get_display

@app.template_filter('rtl')
def rtl_filter(value):
    if value is None:
        return ''
    text = str(value)
    return get_display(arabic_reshaper.reshape(text))

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# ==================== الصفحات الرئيسية ====================
@app.route('/')
def index():
    # إحصائيات عامة
    total_suppliers = Supplier.query.count()
    total_products = Product.query.count()
    total_customers = Customer.query.count()
    frozen_products = Product.query.filter_by(is_frozen=True).count()
    debt_customers = Customer.query.filter(Customer.balance > 0).count()
    total_sales_count = SaleOrder.query.count()

    # مبيعات هذا الشهر
    first_day_of_month = datetime.now().replace(day=1)
    monthly_sales_count = SaleOrder.query.filter(
        SaleOrder.sale_date >= first_day_of_month
    ).count()

    # آخر 5 مشتريات ومبيعات
    recent_purchases = PurchaseOrder.query.order_by(
        PurchaseOrder.order_date.desc()
    ).limit(5).all()

    recent_sales = SaleOrder.query.order_by(
        SaleOrder.sale_date.desc()
    ).limit(5).all()

    return render_template('index.html',
                           total_suppliers=total_suppliers,
                           total_products=total_products,
                           total_customers=total_customers,
                           frozen_products=frozen_products,
                           debt_customers=debt_customers,
                           total_sales_count=total_sales_count,
                           monthly_sales_count=monthly_sales_count,
                           recent_purchases=recent_purchases,
                           recent_sales=recent_sales)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()

        if user and check_password_hash(user.password, password) and user.is_active:
            login_user(user)
            flash(f'مرحباً {user.full_name}', 'success')
            return redirect(url_for('dashboard'))
        flash('اسم المستخدم أو كلمة المرور غير صحيحة', 'danger')
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('تم تسجيل الخروج بنجاح', 'info')
    return redirect(url_for('login'))


@app.route('/dashboard')
@login_required
def dashboard():
    total_products = Product.query.count()
    low_stock = Product.query.filter(Product.quantity < Product.min_quantity).count()
    total_suppliers = Supplier.query.count()
    total_customers = Customer.query.count()
    debt_customers_count = Customer.query.filter(Customer.balance > 0).count()

    today_sales = db.session.query(db.func.sum(SaleOrder.total_amount)).filter(
        db.func.date(SaleOrder.sale_date) == datetime.now().date()
    ).scalar() or 0

    total_deposits = db.session.query(db.func.sum(FreezeDeposit.amount)).filter(
        FreezeDeposit.is_returned == False
    ).scalar() or 0

    # مبيعات هذا الشهر
    first_day_of_month = datetime.now().replace(day=1)
    monthly_sales = db.session.query(db.func.sum(SaleOrder.total_amount)).filter(
        SaleOrder.sale_date >= first_day_of_month
    ).scalar() or 0

    # تحصيلات هذا الشهر
    total_collections_month = db.session.query(db.func.sum(Collection.amount)).filter(
        Collection.collection_date >= first_day_of_month
    ).scalar() or 0

    # المنتجات المضافة هذا الشهر
    products_added_this_month = Product.query.filter(
        Product.created_at >= first_day_of_month
    ).count()

    recent_purchases = PurchaseOrder.query.order_by(PurchaseOrder.order_date.desc()).limit(5).all()
    recent_sales = SaleOrder.query.order_by(SaleOrder.sale_date.desc()).limit(5).all()

    return render_template('dashboard.html',
                           total_products=total_products,
                           low_stock=low_stock,
                           total_suppliers=total_suppliers,
                           total_customers=total_customers,
                           debt_customers_count=debt_customers_count,
                           today_sales=today_sales,
                           total_deposits=total_deposits,
                           monthly_sales=monthly_sales,
                           total_collections_month=total_collections_month,
                           products_added_this_month=products_added_this_month,
                           recent_purchases=recent_purchases,
                           recent_sales=recent_sales)
# ==================== إدارة المستخدمين ====================

@app.route('/users')
@login_required
@permission_required(Permission.MANAGE_USERS)
def users_list():
    users = User.query.all()
    return render_template('users/index.html', users=users)


@app.route('/users/add', methods=['GET', 'POST'])
@login_required
@permission_required(Permission.MANAGE_USERS)
def add_user():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        role = request.form.get('role')
        full_name = request.form.get('full_name')
        phone = request.form.get('phone')
        email = request.form.get('email')

        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            flash('اسم المستخدم موجود بالفعل', 'danger')
            return redirect(url_for('add_user'))

        user = User(
            username=username,
            password=generate_password_hash(password),
            role=role,
            full_name=full_name,
            phone=phone,
            email=email
        )
        db.session.add(user)
        db.session.commit()
        flash('تم إضافة المستخدم بنجاح', 'success')
        return redirect(url_for('users_list'))

    roles = [(r.value, r.name) for r in UserRole]
    return render_template('users/add.html', roles=roles)


@app.route('/users/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required(Permission.MANAGE_USERS)
def edit_user(id):
    user = User.query.get_or_404(id)
    if request.method == 'POST':
        user.username = request.form.get('username')
        user.role = request.form.get('role')
        user.full_name = request.form.get('full_name')
        user.phone = request.form.get('phone')
        user.email = request.form.get('email')
        user.is_active = request.form.get('is_active') == 'on'

        new_password = request.form.get('password')
        if new_password:
            user.password = generate_password_hash(new_password)

        db.session.commit()
        flash('تم تحديث بيانات المستخدم بنجاح', 'success')
        return redirect(url_for('users_list'))

    roles = [(r.value, r.name) for r in UserRole]
    return render_template('users/edit.html', user=user, roles=roles)


@app.route('/users/delete/<int:id>')
@login_required
@permission_required(Permission.MANAGE_USERS)
def delete_user(id):
    user = User.query.get_or_404(id)
    if user.id == current_user.id:
        flash('لا يمكن حذف حسابك الحالي', 'danger')
        return redirect(url_for('users_list'))

    db.session.delete(user)
    db.session.commit()
    flash('تم حذف المستخدم بنجاح', 'success')
    return redirect(url_for('users_list'))


# ==================== إدارة الموظفين ====================

@app.route('/employees')
@login_required
@permission_required(Permission.MANAGE_EMPLOYEES)
def employees():
    employees_list = Employee.query.all()
    return render_template('employees/index.html', employees=employees_list)


@app.route('/employees/add', methods=['GET', 'POST'])
@login_required
@permission_required(Permission.MANAGE_EMPLOYEES)
def add_employee():
    if request.method == 'POST':
        full_name = request.form.get('full_name')
        position = request.form.get('position')
        phone = request.form.get('phone')
        email = request.form.get('email')
        address = request.form.get('address')
        salary = float(request.form.get('salary', 0))
        hire_date_str = request.form.get('hire_date')
        hire_date = datetime.strptime(hire_date_str, '%Y-%m-%d') if hire_date_str else datetime.now()

        employee = Employee(
            full_name=full_name,
            position=position,
            phone=phone,
            email=email,
            address=address,
            salary=salary,
            hire_date=hire_date
        )
        db.session.add(employee)
        db.session.flush()

        create_user = request.form.get('create_user') == 'on'
        if create_user:
            username = request.form.get('username')
            password = request.form.get('password')

            role_map = {
                'مسؤول الصندوق': UserRole.CASHIER.value,
                'أمين المخزن': UserRole.STORE_KEEPER.value,
                'المحصل': UserRole.COLLECTOR.value,
                'مسؤول المشتريات': UserRole.PURCHASE_MANAGER.value,
                'مسؤول المبيعات': UserRole.SALES_MANAGER.value,
                'مدير': UserRole.ADMIN.value
            }
            role = role_map.get(position, 'employee')

            user = User(
                username=username,
                password=generate_password_hash(password),
                role=role,
                full_name=full_name,
                phone=phone,
                email=email
            )
            db.session.add(user)
            db.session.flush()
            employee.user_id = user.id

        db.session.commit()
        flash('تم إضافة الموظف بنجاح', 'success')
        return redirect(url_for('employees'))

    positions = ['مسؤول الصندوق', 'أمين المخزن', 'المحصل', 'مسؤول المشتريات', 'مسؤول المبيعات', 'مدير']
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('employees/add.html', positions=positions, today=today)


@app.route('/employees/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required(Permission.MANAGE_EMPLOYEES)
def edit_employee(id):
    employee = Employee.query.get_or_404(id)
    if request.method == 'POST':
        employee.full_name = request.form.get('full_name')
        employee.position = request.form.get('position')
        employee.phone = request.form.get('phone')
        employee.email = request.form.get('email')
        employee.address = request.form.get('address')
        employee.salary = float(request.form.get('salary', 0))
        employee.is_active = request.form.get('is_active') == 'on'

        db.session.commit()
        flash('تم تحديث بيانات الموظف بنجاح', 'success')
        return redirect(url_for('employees'))

    positions = ['مسؤول الصندوق', 'أمين المخزن', 'المحصل', 'مسؤول المشتريات', 'مسؤول المبيعات', 'مدير']
    return render_template('employees/edit.html', employee=employee, positions=positions)


@app.route('/api/employee/<int:id>/details')
@login_required
@permission_required(Permission.MANAGE_EMPLOYEES)
def employee_details(id):
    """جلب تفاصيل الموظف"""
    employee = Employee.query.get_or_404(id)

    user_data = None
    if employee.user_id:
        user = User.query.get(employee.user_id)
        if user:
            user_data = {
                'username': user.username,
                'role': user.role,
                'is_active': user.is_active
            }

    return jsonify({
        'id': employee.id,
        'full_name': employee.full_name,
        'position': employee.position,
        'phone': employee.phone,
        'email': employee.email,
        'address': employee.address,
        'salary': employee.salary,
        'hire_date': employee.hire_date.strftime('%Y-%m-%d') if employee.hire_date else None,
        'is_active': employee.is_active,
        'user': user_data
    })

@app.route('/employees/delete/<int:id>')
@login_required
@permission_required(Permission.MANAGE_EMPLOYEES)
def delete_employee(id):
    employee = Employee.query.get_or_404(id)
    if employee.user_id:
        user = User.query.get(employee.user_id)
        if user:
            db.session.delete(user)
    db.session.delete(employee)
    db.session.commit()
    flash('تم حذف الموظف بنجاح', 'success')
    return redirect(url_for('employees'))


# ==================== إدارة الموردين ====================

# تعديل صلاحيات الموردين - لمسؤول المشتريات فقط
@app.route('/suppliers')
@login_required
@permission_required(Permission.MANAGE_SUPPLIERS)
def suppliers():

    suppliers_list = Supplier.query.all()

    # تحديث أرصدة الموردين تلقائياً
    for supplier in suppliers_list:

        # إجمالي المشتريات
        purchases_total = db.session.query(
            db.func.coalesce(db.func.sum(PurchaseOrder.total_amount), 0)
        ).filter(
            PurchaseOrder.supplier_id == supplier.id,
            PurchaseOrder.status != 'cancelled'
        ).scalar()

        # إجمالي المدفوعات المعتمدة
        payments_total = db.session.query(
            db.func.coalesce(db.func.sum(SupplierPayment.amount), 0)
        ).filter(
            SupplierPayment.supplier_id == supplier.id,
            SupplierPayment.cash_status == 'approved'
        ).scalar()

        # الرصيد الحقيقي
        supplier.balance = float(purchases_total or 0) - float(payments_total or 0)

    db.session.commit()

    return render_template(
        'suppliers/index.html',
        suppliers=suppliers_list
    )


@app.route('/suppliers/add', methods=['GET', 'POST'])
@login_required
@permission_required(Permission.MANAGE_SUPPLIERS)
def add_supplier():
    from accounting_utils import get_account_info_by_code, get_account_code_by_id
    if request.method == 'POST':
        name = request.form.get('name')
        contact_person = request.form.get('contact_person')
        phone = request.form.get('phone')
        email = request.form.get('email')
        address = request.form.get('address')

        # إنشاء حساب فرعي للمورد
        # تحديد كود الحساب الفرعي: نبحث عن آخر حساب من نفس النوع
        last_supplier_account = FinancialAccount.query.filter(
            FinancialAccount.account_code.like('2001%'),
            FinancialAccount.account_code != '2001'
        ).order_by(FinancialAccount.account_code.desc()).first()

        if last_supplier_account:
            last_code = int(last_supplier_account.account_code)
            new_code = str(last_code + 1)
        else:
            new_code = '20010001'

        new_account = FinancialAccount(
            account_code=new_code,
            account_name=f'مورد: {name}',
            account_type='liability',
            balance=0,
            is_active=True
        )
        db.session.add(new_account)
        db.session.flush()

        supplier = Supplier(
            name=name,
            contact_person=contact_person,
            phone=phone,
            email=email,
            address=address,
            account_id=new_account.id
        )
        db.session.add(supplier)
        db.session.commit()
        flash(f'تم إضافة المورد {name} مع حساب فرعي {new_code}', 'success')
        return redirect(url_for('suppliers'))
    return render_template('suppliers/add.html')

@app.route('/suppliers/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required(Permission.MANAGE_SUPPLIERS)
def edit_supplier(id):
    supplier = Supplier.query.get_or_404(id)
    if request.method == 'POST':
        supplier.name = request.form.get('name')
        supplier.contact_person = request.form.get('contact_person')
        supplier.phone = request.form.get('phone')
        supplier.email = request.form.get('email')
        supplier.address = request.form.get('address')
        db.session.commit()
        flash('تم تحديث بيانات المورد بنجاح', 'success')
        return redirect(url_for('suppliers'))
    return render_template('suppliers/edit.html', supplier=supplier)


@app.route('/suppliers/delete/<int:id>')
@login_required
@permission_required(Permission.MANAGE_SUPPLIERS)
def delete_supplier(id):
    supplier = Supplier.query.get_or_404(id)
    db.session.delete(supplier)
    db.session.commit()
    flash('تم حذف المورد بنجاح', 'success')
    return redirect(url_for('suppliers'))


# ==================== إدارة العملاء ====================

# تعديل صلاحيات العملاء - لمسؤول المبيعات والمحصل
@app.route('/customers')
@login_required
@permission_required(Permission.MANAGE_CUSTOMERS)
def customers():
    customers_list = Customer.query.all()
    return render_template('customers/index.html', customers=customers_list)


@app.route('/customers/add', methods=['GET', 'POST'])
@login_required
@permission_required(Permission.MANAGE_CUSTOMERS)
def add_customer():
    if request.method == 'POST':
        name = request.form.get('name')
        type_ = request.form.get('type')
        phone = request.form.get('phone')
        address = request.form.get('address')
        credit_limit = float(request.form.get('credit_limit', 5000))

        # إنشاء حساب فرعي للعميل
        last_customer_account = FinancialAccount.query.filter(
            FinancialAccount.account_code.like('1004%'),
            FinancialAccount.account_code != '1004'
        ).order_by(FinancialAccount.account_code.desc()).first()

        if last_customer_account:
            last_code = int(last_customer_account.account_code)
            new_code = str(last_code + 1)
        else:
            new_code = '10040001'

        new_account = FinancialAccount(
            account_code=new_code,
            account_name=f'عميل: {name}',
            account_type='asset',
            balance=0,
            is_active=True
        )
        db.session.add(new_account)
        db.session.flush()

        customer = Customer(
            name=name,
            type=type_,
            phone=phone,
            address=address,
            credit_limit=credit_limit,
            account_id=new_account.id
        )
        db.session.add(customer)
        db.session.commit()
        flash(f'تم إضافة العميل {name} مع حساب فرعي {new_code}', 'success')
        return redirect(url_for('customers'))
    return render_template('customers/add.html')

@app.route('/customers/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required(Permission.MANAGE_CUSTOMERS)
def edit_customer(id):
    customer = Customer.query.get_or_404(id)
    if request.method == 'POST':
        customer.name = request.form.get('name')
        customer.type = request.form.get('type')
        customer.phone = request.form.get('phone')
        customer.address = request.form.get('address')
        customer.credit_limit = float(request.form.get('credit_limit', 5000))
        db.session.commit()
        flash('تم تحديث بيانات العميل بنجاح', 'success')
        return redirect(url_for('customers'))
    return render_template('customers/edit.html', customer=customer)


# ==================== إدارة المخزون (عرض فقط) ====================
# ==================== إدارة المنتجات ====================

@app.route('/products')
@login_required
@permission_required(Permission.VIEW_INVENTORY)
def products_list():
    """عرض قائمة المنتجات"""
    products = Product.query.order_by(Product.name).all()

    # إحصائيات
    total_products = len(products)
    total_quantity = sum(p.quantity for p in products)
    total_value = sum(p.quantity * p.purchase_price for p in products)
    low_stock_count = sum(1 for p in products if p.quantity < p.min_quantity)
    frozen_count = sum(1 for p in products if p.is_frozen)

    return render_template('products/index.html',
                           products=products,
                           total_products=total_products,
                           total_quantity=total_quantity,
                           total_value=total_value,
                           low_stock_count=low_stock_count,
                           frozen_count=frozen_count)


@app.route('/products/add', methods=['GET', 'POST'])
@login_required
@permission_required(Permission.EDIT_PRODUCT)
def add_product():
    """إضافة منتج جديد"""
    if request.method == 'POST':
        name = request.form.get('name')
        category = request.form.get('category')
        unit = request.form.get('unit')
        purchase_price = float(request.form.get('purchase_price', 0))
        selling_price = float(request.form.get('selling_price', 0))
        quantity = int(request.form.get('quantity', 0))
        min_quantity = int(request.form.get('min_quantity', 10))
        location = request.form.get('location')
        is_frozen = request.form.get('is_frozen') == 'on'
        freeze_deposit = float(request.form.get('freeze_deposit', 0))

        # التحقق من عدم وجود منتج بنفس الاسم
        existing = Product.query.filter_by(name=name).first()
        if existing:
            flash('منتج بنفس الاسم موجود بالفعل', 'danger')
            return redirect(url_for('add_product'))

        product = Product(
            name=name,
            category=category,
            unit=unit,
            purchase_price=purchase_price,
            selling_price=selling_price,
            quantity=quantity,
            min_quantity=min_quantity,
            location=location,
            is_frozen=is_frozen,
            freeze_deposit=freeze_deposit
        )
        db.session.add(product)
        db.session.commit()

        flash(f'تم إضافة المنتج {name} بنجاح', 'success')
        return redirect(url_for('products_list'))

    categories = ['مياه', 'مشروبات غازية', 'مشروبات طاقة', 'أجبان', 'ألبان', 'زيوت', 'مواد غذائية', 'مجمدة']
    return render_template('products/add.html', categories=categories)


@app.route('/products/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required(Permission.EDIT_PRODUCT)
def edit_product_standalone(id):
    """تعديل منتج"""
    product = Product.query.get_or_404(id)

    if request.method == 'POST':
        product.name = request.form.get('name')
        product.category = request.form.get('category')
        product.unit = request.form.get('unit')
        product.purchase_price = float(request.form.get('purchase_price', 0))
        product.selling_price = float(request.form.get('selling_price', 0))
        product.min_quantity = int(request.form.get('min_quantity', 10))
        product.location = request.form.get('location')
        product.is_frozen = request.form.get('is_frozen') == 'on'
        product.freeze_deposit = float(request.form.get('freeze_deposit', 0))

        db.session.commit()
        flash('تم تحديث بيانات المنتج بنجاح', 'success')
        return redirect(url_for('products_list'))

    categories = ['مياه', 'مشروبات غازية', 'مشروبات طاقة', 'أجبان', 'ألبان', 'زيوت', 'مواد غذائية', 'مجمدة']
    return render_template('products/edit.html', product=product, categories=categories)


@app.route('/products/delete/<int:id>')
@login_required
@permission_required(Permission.EDIT_PRODUCT)
def delete_product(id):
    """حذف منتج"""
    product = Product.query.get_or_404(id)

    # التحقق من وجود مشتريات مرتبطة
    purchases = PurchaseItem.query.filter_by(product_id=id).count()
    if purchases > 0:
        flash('لا يمكن حذف منتج له مشتريات مرتبطة', 'danger')
        return redirect(url_for('products_list'))

    db.session.delete(product)
    db.session.commit()
    flash('تم حذف المنتج بنجاح', 'success')
    return redirect(url_for('products_list'))


@app.route('/api/products/search_json')
@login_required
def search_products_json():
    """API للبحث عن المنتجات (للاستخدام في المشتريات)"""
    query = request.args.get('q', '')
    products = Product.query.filter(Product.name.contains(query)).limit(20).all()
    return jsonify([{
        'id': p.id,
        'name': p.name,
        'category': p.category,
        'purchase_price': p.purchase_price,
        'selling_price': p.selling_price,
        'quantity': p.quantity,
        'unit': p.unit
    } for p in products])

@app.route('/inventory')
@login_required
@permission_required(Permission.VIEW_INVENTORY)
def inventory():
    products = Product.query.all()
    return render_template('inventory/index.html', products=products)


@app.route('/inventory/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required(Permission.EDIT_PRODUCT)
def edit_product(id):
    product = Product.query.get_or_404(id)
    if request.method == 'POST':
        product.name = request.form.get('name')
        product.category = request.form.get('category')
        product.unit = request.form.get('unit')
        product.selling_price = float(request.form.get('selling_price', 0))
        product.min_quantity = int(request.form.get('min_quantity', 10))
        product.location = request.form.get('location')
        product.is_frozen = request.form.get('is_frozen') == 'on'
        product.freeze_deposit = float(request.form.get('freeze_deposit', 0))
        db.session.commit()
        flash('تم تحديث بيانات المنتج بنجاح', 'success')
        return redirect(url_for('inventory'))
    return render_template('inventory/edit.html', product=product)


# ==================== إدارة المشتريات ====================


from sqlalchemy.orm import selectinload


@app.route('/purchases')
@app.route('/purchases/page/<int:page>')
@login_required
@permission_required(Permission.VIEW_PURCHASES)
def purchases(page=1):
    """عرض قائمة المشتريات مع دعم الترقيم"""
    from sqlalchemy import func
    from sqlalchemy import case

    per_page = 10

    # الحصول على المشتريات مع الترقيم
    pagination = PurchaseOrder.query.order_by(
        PurchaseOrder.order_date.desc(),
        PurchaseOrder.id.desc()
    ).paginate(page=page, per_page=per_page, error_out=False)

    # ✅ حساب إجمالي المدفوعات من جميع الفواتير
    # للفواتير النقدية: نأخذ paid_amount
    # للفواتير الآجلة: نجمع من supplier_payments

    # مدفوعات الفواتير النقدية
    cash_paid = db.session.query(
        func.coalesce(func.sum(PurchaseOrder.paid_amount).filter(
            PurchaseOrder.payment_type == 'cash'
        ), 0)
    ).scalar() or 0

    # مدفوعات الفواتير الآجلة (من جدول supplier_payments)
    credit_paid = db.session.query(
        func.coalesce(func.sum(SupplierPayment.amount).filter(
            SupplierPayment.cash_status == 'approved'
        ), 0)
    ).scalar() or 0

    total_paid = cash_paid + credit_paid

    # إجمالي المشتريات
    total_amount = db.session.query(func.sum(PurchaseOrder.total_amount)).scalar() or 0

    # عدد الفواتير
    total_count = PurchaseOrder.query.count()

    return render_template('purchases/index.html',
                           purchases=pagination,
                           total_amount=total_amount,
                           total_paid=total_paid,
                           total_count=total_count)

# purchases
@app.route('/purchases/add', methods=['GET', 'POST'])
@login_required
@permission_required(Permission.ADD_PURCHASE)
def add_purchase():
    from accounting_utils import create_double_entry

    suppliers = Supplier.query.all()
    existing_products = Product.query.all()

    if request.method == 'POST':
        try:
            supplier_id = request.form.get('supplier_id')
            payment_type = request.form.get('payment_type')
            total_amount = float(request.form.get('total_amount') or 0)
            paid_amount = float(request.form.get('paid_amount') or 0)

            supplier = Supplier.query.get(supplier_id)
            if not supplier:
                flash('المورد غير موجود', 'danger')
                return redirect(url_for('add_purchase'))

            items_data = json.loads(request.form.get('items_data', '[]'))
            if not items_data:
                flash('يجب إضافة صنف واحد على الأقل', 'danger')
                return redirect(url_for('add_purchase'))

            # شراء نقدي
            if payment_type == 'cash':
                purchase = PurchaseOrder(
                    supplier_id=supplier_id,
                    total_amount=total_amount,
                    paid_amount=0,
                    payment_type='cash',
                    status='pending',
                    cash_status='pending',
                    created_by=current_user.id
                )

                db.session.add(purchase)
                db.session.flush()

                for item in items_data:
                    quantity = float(item['quantity'])
                    unit_price = float(item['unit_price'])

                    db.session.add(PurchaseItem(
                        purchase_order_id=purchase.id,
                        product_id=item['product_id'],
                        quantity=quantity,
                        unit_price=unit_price,
                        total_price=quantity * unit_price
                    ))

                db.session.commit()
                flash(f'تم تسجيل عملية شراء نقدية بمبلغ {total_amount:,.2f} ريال وتنتظر موافقة الصندوق', 'info')

            # شراء آجل
            else:
                # شراء آجل مع دفعة مقدمة
                if paid_amount > 0 and paid_amount < total_amount:
                    purchase = PurchaseOrder(
                        supplier_id=supplier_id,
                        total_amount=total_amount,
                        paid_amount=paid_amount,
                        payment_type='credit',
                        status='pending',
                        cash_status='pending',
                        created_by=current_user.id
                    )

                    db.session.add(purchase)
                    db.session.flush()

                    for item in items_data:
                        quantity = float(item['quantity'])
                        unit_price = float(item['unit_price'])

                        db.session.add(PurchaseItem(
                            purchase_order_id=purchase.id,
                            product_id=item['product_id'],
                            quantity=quantity,
                            unit_price=unit_price,
                            total_price=quantity * unit_price
                        ))

                    # المديونية المتبقية فقط
                    supplier.balance = float(supplier.balance or 0) + (total_amount - paid_amount)

                    if supplier.account_id:
                        supplier_account = db.session.get(FinancialAccount, supplier.account_id)
                        if supplier_account:
                            supplier_account.balance = float(supplier_account.balance or 0) + (total_amount - paid_amount)

                    db.session.commit()
                    flash(f'تم تسجيل شراء آجل مع دفعة مقدمة {paid_amount:,.2f} ريال وتنتظر موافقة الصندوق', 'info')

                # شراء آجل بدون دفعة
                else:
                    purchase = PurchaseOrder(
                        supplier_id=supplier_id,
                        total_amount=total_amount,
                        paid_amount=0,
                        payment_type='credit',
                        status='completed',
                        cash_status='approved',
                        cash_approved_by=current_user.id,
                        cash_approved_at=datetime.now(),
                        created_by=current_user.id
                    )

                    db.session.add(purchase)
                    db.session.flush()

                    for item in items_data:
                        quantity = float(item['quantity'])
                        unit_price = float(item['unit_price'])

                        db.session.add(PurchaseItem(
                            purchase_order_id=purchase.id,
                            product_id=item['product_id'],
                            quantity=quantity,
                            unit_price=unit_price,
                            total_price=quantity * unit_price
                        ))

                        product = db.session.get(Product, item['product_id'])
                        if product:
                            product.quantity += quantity
                            product.purchase_price = unit_price

                    supplier.balance = float(supplier.balance or 0) + total_amount

                    if supplier.account_id:
                        supplier_account = db.session.get(FinancialAccount, supplier.account_id)
                        if supplier_account:
                            supplier_account.balance = float(supplier_account.balance or 0) + total_amount

                    db.session.commit()

                    create_supplier_double_entry(
                        supplier.id,
                        '1003',  # المخزون
                        total_amount,
                        'PUR',
                        purchase.id,
                        f'شراء آجل من المورد {supplier.name}',
                        current_user.id
                    )

                    flash('تم تسجيل عملية الشراء الآجلة وتحديث المخزون ومديونية المورد بنجاح', 'success')

            return redirect(url_for('purchases'))

        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء حفظ عملية الشراء: {str(e)}', 'danger')
            return redirect(url_for('add_purchase'))

    return render_template('purchases/add.html', suppliers=suppliers, products=existing_products)

@app.route('/purchases/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required(Permission.ADD_PURCHASE)
def edit_purchase(id):
    purchase = PurchaseOrder.query.get_or_404(id)
    suppliers = Supplier.query.all()
    products = Product.query.all()

    # التحقق: لا يمكن تعديل عملية مكتملة
    if purchase.status == 'completed':
        flash('لا يمكن تعديل عملية شراء مكتملة', 'danger')
        return redirect(url_for('purchases'))

    if purchase.cash_status == 'approved':
        flash('لا يمكن تعديل عملية شراء تمت الموافقة عليها', 'danger')
        return redirect(url_for('purchases'))

    if request.method == 'POST':
        supplier_id = request.form.get('supplier_id')
        payment_type = request.form.get('payment_type')
        total_amount = float(request.form.get('total_amount'))
        paid_amount = float(request.form.get('paid_amount', 0))

        # تحديث بيانات أمر الشراء
        purchase.supplier_id = supplier_id
        purchase.total_amount = total_amount
        purchase.paid_amount = paid_amount
        purchase.payment_type = payment_type

        # حذف الأصناف القديمة
        for item in purchase.items:
            db.session.delete(item)

        db.session.flush()

        # إضافة الأصناف الجديدة
        items_data = json.loads(request.form.get('items_data', '[]'))
        for item in items_data:
            purchase_item = PurchaseItem(
                purchase_order_id=purchase.id,
                product_id=item['product_id'],
                quantity=int(item['quantity']),
                unit_price=float(item['unit_price']),
                total_price=int(item['quantity']) * float(item['unit_price'])
            )
            db.session.add(purchase_item)

        # تحديث حالة العملية
        if payment_type == 'cash':
            purchase.status = 'pending'
            purchase.cash_status = 'pending'
        else:
            purchase.status = 'completed'
            purchase.cash_status = 'approved'

        db.session.commit()
        flash('تم تعديل عملية الشراء بنجاح', 'success')
        return redirect(url_for('purchases'))

    # عرض صفحة التعديل - مع التحقق من وجود المنتج
    items_json = []
    for item in purchase.items:
        product_name = item.product.name if item.product else 'منتج غير موجود'
        items_json.append({
            'product_id': item.product_id,
            'product_name': product_name,
            'quantity': item.quantity,
            'unit_price': item.unit_price,
            'total_price': item.total_price
        })

    items_json_str = json.dumps(items_json)

    return render_template('purchases/edit.html',
                           purchase=purchase,
                           suppliers=suppliers,
                           products=products,
                           items_json=items_json_str)
@app.route('/purchases/approve/<int:id>', methods=['POST'])
@login_required
@permission_required(Permission.MANAGE_CASH)
def approve_purchase_direct(id):
    from accounting_utils import create_double_entry

    purchase = PurchaseOrder.query.get_or_404(id)

    if purchase.cash_status == 'approved':
        flash('تمت الموافقة على هذه العملية مسبقاً', 'warning')
        return redirect(url_for('purchases'))

    purchase.cash_status = 'approved'
    purchase.cash_approved_by = current_user.id
    purchase.cash_approved_at = datetime.now()
    purchase.status = 'completed'

    if purchase.payment_type == 'cash':
        purchase.paid_amount = purchase.total_amount

    for item in purchase.items:
        product = db.session.get(Product, item.product_id)
        if product:
            product.quantity += item.quantity
            product.purchase_price = item.unit_price

    remaining = float(purchase.total_amount or 0) - float(purchase.paid_amount or 0)
    if remaining > 0:
        supplier = db.session.get(Supplier, purchase.supplier_id)
        if supplier:
            supplier.balance = float(supplier.balance or 0) + remaining

            if supplier.account_id:
                supplier_account = db.session.get(FinancialAccount, supplier.account_id)
                if supplier_account:
                    supplier_account.balance = float(supplier_account.balance or 0) + remaining

    db.session.commit()

    if purchase.paid_amount > 0:
        if purchase.payment_type == 'cash':
            create_double_entry(
                '1003',          # المخزون
                '1001',          # الصندوق
                purchase.paid_amount,
                'PUR',
                purchase.id,
                f'شراء نقدي من مورد {purchase.supplier.name}',
                current_user.id
            )
        else:
            create_supplier_double_entry(
                purchase.supplier_id,
                '1003',
                purchase.paid_amount,
                'PUR',
                purchase.id,
                f'دفعة مقدمة لمورد {purchase.supplier.name}',
                current_user.id
            )

    flash(
        f'تمت الموافقة على عملية الشراء رقم {purchase.id}',
        'success'
    )
    return redirect(url_for('purchases'))

@app.route('/purchases/delete/<int:id>')
@login_required
@permission_required(Permission.ADD_PURCHASE)
def delete_purchase(id):
    purchase = PurchaseOrder.query.get_or_404(id)

    # التحقق: لا يمكن حذف عملية مكتملة أو موافق عليها
    if purchase.cash_status == 'approved':
        flash('لا يمكن حذف عملية شراء تمت الموافقة عليها مسبقاً', 'danger')
        return redirect(url_for('purchases'))

    if purchase.status == 'completed':
        flash('لا يمكن حذف عملية شراء مكتملة', 'danger')
        return redirect(url_for('purchases'))

    # حذف عناصر الشراء أولاً
    for item in purchase.items:
        db.session.delete(item)

    # حذف أمر الشراء
    db.session.delete(purchase)
    db.session.commit()

    flash('تم حذف عملية الشراء بنجاح', 'success')
    return redirect(url_for('purchases'))


@app.route('/purchases/cancel/<int:id>')
@login_required
@permission_required(Permission.ADD_PURCHASE)
def cancel_purchase(id):
    """إلغاء عملية شراء (بدلاً من حذفها)"""
    purchase = PurchaseOrder.query.get_or_404(id)

    # التحقق: لا يمكن إلغاء عملية مكتملة
    if purchase.status == 'completed':
        flash('لا يمكن إلغاء عملية شراء مكتملة', 'danger')
        return redirect(url_for('purchases'))

    # إذا كانت العملية قد تمت الموافقة عليها مسبقاً، نعيد المخزون
    if purchase.cash_status == 'approved':
        for item in purchase.items:
            product = Product.query.get(item.product_id)
            if product:
                product.quantity -= item.quantity

    # تحديث الحالة إلى ملغاة
    purchase.status = 'cancelled'
    purchase.cash_status = 'rejected'
    purchase.cash_rejection_reason = 'تم الإلغاء بواسطة المستخدم'

    db.session.commit()
    flash('تم إلغاء عملية الشراء بنجاح', 'success')
    return redirect(url_for('purchases'))


@app.route('/purchases/print/<int:id>')
@login_required
@permission_required(Permission.VIEW_PURCHASES)
def print_purchase(id):
    """طباعة تفاصيل عملية الشراء"""
    purchase = PurchaseOrder.query.get_or_404(id)
    return render_template('purchases/print.html', purchase=purchase)


# ==================== إدارة سداد الموردين ====================

@app.route('/suppliers/payments')
@login_required
@permission_required(Permission.MANAGE_SUPPLIERS)
def supplier_payments():
    """عرض سجل سداد الموردين"""
    suppliers = Supplier.query.filter(Supplier.balance > 0).all()
    payments = SupplierPayment.query.order_by(SupplierPayment.payment_date.desc()).all()

    total_paid = sum(p.amount for p in payments if p.cash_status == 'approved')
    pending_payments = SupplierPayment.query.filter_by(cash_status='pending').count()

    # حساب الفواتير المفتوحة لكل مورد
    for supplier in suppliers:
        open_invoices = PurchaseOrder.query.filter(
            PurchaseOrder.supplier_id == supplier.id,
            PurchaseOrder.total_amount > PurchaseOrder.paid_amount
        ).all()
        supplier.open_invoices = open_invoices
        supplier.open_invoices_count = len(open_invoices)

    return render_template('suppliers/payments.html',
                           suppliers=suppliers,
                           payments=payments,
                           total_paid=total_paid,
                           pending_payments=pending_payments,
                           Permission=Permission)   # ✅ هذا السطر يمنع خطأ 'Permission' is undefined

@app.route('/suppliers/payments/add', methods=['POST'])
@login_required
@permission_required(Permission.MANAGE_SUPPLIERS)
def add_supplier_payment():
    supplier_id = request.form.get('supplier_id')
    purchase_order_id = request.form.get('purchase_order_id')
    amount = float(request.form.get('amount', 0))
    payment_type = request.form.get('payment_type', 'cash')
    reference_number = request.form.get('reference_number', '')
    notes = request.form.get('notes', '')

    supplier = Supplier.query.get(supplier_id)
    if not supplier:
        flash('المورد غير موجود', 'danger')
        return redirect(url_for('supplier_payments'))

    if amount <= 0:
        flash('المبلغ يجب أن يكون أكبر من صفر', 'danger')
        return redirect(url_for('supplier_payments'))

    if amount > supplier.balance:
        flash(f'المبلغ المطلوب ({amount:,.2f} ريال) يتجاوز المديونية ({supplier.balance:,.2f} ريال)', 'danger')
        return redirect(url_for('supplier_payments'))

    if not purchase_order_id:
        purchase_order = PurchaseOrder.query.filter(
            PurchaseOrder.supplier_id == supplier_id,
            PurchaseOrder.total_amount > PurchaseOrder.paid_amount
        ).order_by(PurchaseOrder.order_date.desc()).first()

        if purchase_order:
            purchase_order_id = purchase_order.id

    payment = SupplierPayment(
        supplier_id=supplier_id,
        purchase_order_id=purchase_order_id if purchase_order_id else None,
        amount=amount,
        payment_type=payment_type,
        reference_number=reference_number,
        notes=notes,
        created_by=current_user.id,
        cash_status='pending'
    )
    db.session.add(payment)
    db.session.commit()

    flash(f'تم تسجيل دفعة للمورد {supplier.name} بمبلغ {amount:,.2f} ريال وتنتظر موافقة الصندوق', 'success')
    return redirect(url_for('supplier_payments'))

@app.route('/suppliers/payments/approve/<int:id>', methods=['POST'])
@login_required
@permission_required(Permission.MANAGE_CASH)
def approve_supplier_payment(id):
    from accounting_utils import create_supplier_payment_entry

    payment = SupplierPayment.query.get_or_404(id)

    if payment.cash_status == 'approved':
        flash('تمت الموافقة على هذه الدفعة مسبقاً', 'warning')
        return redirect(url_for('supplier_payments'))

    supplier = payment.supplier
    if not supplier:
        flash('المورد غير موجود', 'danger')
        return redirect(url_for('supplier_payments'))

    try:
        payment.cash_status = 'approved'
        payment.cash_approved_by = current_user.id
        payment.cash_approved_at = datetime.utcnow()

        # تحديث رصيد المورد في جدول suppliers
        supplier.balance = float(supplier.balance or 0) - float(payment.amount or 0)

        # تحديث رصيد الحساب المالي للمورد
        if supplier.account_id:
            account = db.session.get(FinancialAccount, supplier.account_id)
            if account:
                account.balance = float(account.balance or 0) - float(payment.amount or 0)

        # تحديث الفاتورة المرتبطة إن وجدت
        if payment.purchase_order_id:
            purchase = PurchaseOrder.query.get(payment.purchase_order_id)
            if purchase:
                purchase.paid_amount = float(purchase.paid_amount or 0) + float(payment.amount or 0)
                # المبلغ المتبقي يحسب تلقائياً من property
                purchase.status = 'completed' if purchase.remaining_amount <= 0 else 'partial'

        db.session.flush()

        # ✅ إنشاء القيد باستخدام حساب المورد الفرعي
        ok = create_supplier_payment_entry(
            supplier_id=supplier.id,
            amount=float(payment.amount or 0),
            reference_type='PAY',
            reference_id=payment.id,
            description=f'سداد للمورد {supplier.name}',
            user_id=current_user.id
        )

        if not ok:
            raise Exception('فشل إنشاء القيد المحاسبي')

        # تسجيل حركة الصندوق
        db.session.add(CashTransaction(
            type='expense',
            amount=payment.amount,
            description=f'سداد للمورد {supplier.name}',
            reference_type='supplier_payment',
            reference_id=payment.id,
            user_id=current_user.id
        ))

        # تحديث رصيد الصندوق
        cash_box = CashBox.query.first()
        if cash_box:
            cash_box.balance -= payment.amount
            cash_box.updated_at = datetime.now()

        db.session.commit()
        flash(f'تمت الموافقة على سداد {payment.amount:,.2f} ريال للمورد {supplier.name}', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء اعتماد السداد: {str(e)}', 'danger')

    return redirect(url_for('supplier_payments'))

@app.route('/suppliers/payments/reject/<int:id>', methods=['POST'])
@login_required
@permission_required(Permission.MANAGE_CASH)
def reject_supplier_payment(id):
    """رفض سداد مورد"""
    payment = SupplierPayment.query.get_or_404(id)
    reason = request.form.get('reason', '')

    payment.cash_status = 'rejected'
    payment.notes = f"{payment.notes}\nالرفض: {reason}" if payment.notes else f"الرفض: {reason}"

    db.session.commit()
    flash(f'تم رفض سداد المورد: {reason}', 'warning')
    return redirect(url_for('supplier_payments'))

from models import db, User, SaleOrder, SaleItem
from sqlalchemy import func
from sqlalchemy.orm import selectinload


@app.route('/sales')
@login_required
@permission_required(Permission.VIEW_SALES)
def sales():
    """عرض قائمة المبيعات"""
    page = request.args.get('page', 1, type=int)
    per_page = 10

    # الحصول على البيانات مع الترقيم
    pagination = SaleOrder.query.order_by(
        SaleOrder.sale_date.desc()
    ).paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )

    sales_list = pagination.items

    # إحصائيات عامة (من جميع البيانات وليس فقط الصفحة الحالية)
    all_sales = SaleOrder.query.all()
    total_amount = sum(s.total_amount for s in all_sales)
    total_paid = sum(s.paid_amount for s in all_sales)
    total_count = len(all_sales)

    return render_template(
        'sales/index.html',
        sales=sales_list,
        pagination=pagination,  # ✅ تمرير pagination للقالب
        total_amount=total_amount,
        total_paid=total_paid,
        total_count=total_count
    )
@app.route('/sales/add', methods=['GET', 'POST'])
@login_required
@permission_required(Permission.ADD_SALE)
def add_sale():
    from accounting_utils import create_double_entry

    customers = Customer.query.all()
    products = Product.query.filter(Product.quantity > 0).all()

    if request.method == 'POST':
        customer_id = request.form.get('customer_id')
        payment_type = request.form.get('payment_type')
        total_amount = float(request.form.get('total_amount'))
        paid_amount = float(request.form.get('paid_amount', 0))

        customer = Customer.query.get(customer_id)
        if not customer:
            flash('العميل غير موجود', 'danger')
            return redirect(url_for('sales'))

        if payment_type == 'cash':
            sale = SaleOrder(
                customer_id=customer_id,
                total_amount=total_amount,
                paid_amount=0,
                payment_type='cash',
                status='pending',
                cash_status='pending',
                created_by=current_user.id
            )
            db.session.add(sale)
            db.session.flush()

            items_data = json.loads(request.form.get('items_data', '[]'))
            for item in items_data:
                db.session.add(SaleItem(
                    sale_order_id=sale.id,
                    product_id=item['product_id'],
                    quantity=item['quantity'],
                    unit_price=item['unit_price'],
                    total_price=item['quantity'] * item['unit_price']
                ))

            db.session.commit()
            flash('تم تسجيل عملية البيع النقدية وتنتظر موافقة الصندوق', 'info')

        else:
            sale = SaleOrder(
                customer_id=customer_id,
                total_amount=total_amount,
                paid_amount=paid_amount,
                payment_type='credit',
                status='completed',
                cash_status='approved',
                cash_approved_by=current_user.id,
                cash_approved_at=datetime.now(),
                created_by=current_user.id
            )
            db.session.add(sale)
            db.session.flush()

            items_data = json.loads(request.form.get('items_data', '[]'))
            for item in items_data:
                db.session.add(SaleItem(
                    sale_order_id=sale.id,
                    product_id=item['product_id'],
                    quantity=item['quantity'],
                    unit_price=item['unit_price'],
                    total_price=item['quantity'] * item['unit_price']
                ))

                product = Product.query.get(item['product_id'])
                if product:
                    product.quantity -= item['quantity']

            if total_amount > paid_amount:
                customer.balance += (total_amount - paid_amount)

            db.session.commit()

            create_customer_double_entry(
                customer.id,
                '4001',  # إيرادات المبيعات
                total_amount,
                'SAL',
                sale.id,
                f'بيع آجل للعميل {customer.name}',
                current_user.id
            )

            flash('تم تسجيل عملية البيع الآجلة وتحديث المخزون والمبيعات بنجاح', 'success')

        return redirect(url_for('sales'))

    return render_template('sales/add.html', customers=customers, products=products)

# ==================== تعديل وحذف المبيعات ====================

@app.route('/sales/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required(Permission.ADD_SALE)
def edit_sale(id):
    sale = SaleOrder.query.get_or_404(id)
    customers = Customer.query.all()
    products = Product.query.all()

    # التحقق: لا يمكن تعديل عملية مكتملة
    if sale.status == 'completed':
        flash('لا يمكن تعديل عملية بيع مكتملة', 'danger')
        return redirect(url_for('sales'))

    if sale.cash_status == 'approved':
        flash('لا يمكن تعديل عملية بيع تمت الموافقة عليها', 'danger')
        return redirect(url_for('sales'))

    if request.method == 'POST':
        customer_id = request.form.get('customer_id')
        payment_type = request.form.get('payment_type')
        total_amount = float(request.form.get('total_amount'))
        paid_amount = float(request.form.get('paid_amount', 0))

        # حفظ البيانات القديمة للمخزون
        old_items = {item.product_id: item.quantity for item in sale.items}

        # تحديث بيانات أمر البيع
        sale.customer_id = customer_id
        sale.total_amount = total_amount
        sale.paid_amount = paid_amount
        sale.payment_type = payment_type

        # حذف الأصناف القديمة
        for item in sale.items:
            db.session.delete(item)

        db.session.flush()

        # إضافة الأصناف الجديدة
        items_data = json.loads(request.form.get('items_data', '[]'))
        for item in items_data:
            sale_item = SaleItem(
                sale_order_id=sale.id,
                product_id=item['product_id'],
                quantity=int(item['quantity']),
                unit_price=float(item['unit_price']),
                total_price=int(item['quantity']) * float(item['unit_price'])
            )
            db.session.add(sale_item)

        # إذا كانت العملية موافق عليها، تحديث المخزون
        if sale.cash_status == 'approved':
            # إعادة الكميات القديمة
            for product_id, old_qty in old_items.items():
                product = Product.query.get(product_id)
                if product:
                    product.quantity += old_qty

            # خصم الكميات الجديدة
            for item in items_data:
                product = Product.query.get(item['product_id'])
                if product:
                    product.quantity -= item['quantity']

        # تحديث حالة العملية
        if payment_type == 'cash':
            sale.status = 'pending'
            sale.cash_status = 'pending'
        else:
            sale.status = 'completed'
            sale.cash_status = 'approved'

        db.session.commit()
        flash('تم تعديل عملية البيع بنجاح', 'success')
        return redirect(url_for('sales'))

    # عرض صفحة التعديل
    items_json = []
    for item in sale.items:
        product_name = item.product.name if item.product else 'منتج غير موجود'
        items_json.append({
            'product_id': item.product_id,
            'product_name': product_name,
            'quantity': item.quantity,
            'unit_price': item.unit_price,
            'total_price': item.total_price
        })

    items_json_str = json.dumps(items_json)

    return render_template('sales/edit.html',
                           sale=sale,
                           customers=customers,
                           products=products,
                           items_json=items_json_str)

@app.route('/sales/approve/<int:id>', methods=['POST'])
@login_required
@permission_required(Permission.MANAGE_CASH)
def approve_sale_direct(id):
    from accounting_utils import create_double_entry

    sale = SaleOrder.query.get_or_404(id)

    if sale.cash_status == 'approved':
        flash('تمت الموافقة على هذه العملية مسبقاً', 'warning')
        return redirect(url_for('sales'))

    if sale.payment_type != 'cash':
        flash('المبيعات الآجلة لا تحتاج موافقة الصندوق', 'warning')
        return redirect(url_for('sales'))

    sale.cash_status = 'approved'
    sale.cash_approved_by = current_user.id
    sale.cash_approved_at = datetime.now()
    sale.status = 'completed'
    sale.paid_amount = sale.total_amount

    for item in sale.items:
        product = Product.query.get(item.product_id)
        if product:
            product.quantity -= item.quantity

    cash_transaction = CashTransaction(
        type='income',
        amount=sale.total_amount,
        description=f'بيع نقدي من عميل {sale.customer.name}',
        reference_type='sale',
        reference_id=sale.id,
        user_id=current_user.id
    )
    db.session.add(cash_transaction)

    cash_box = CashBox.query.first()
    if cash_box:
        cash_box.balance += sale.total_amount
        cash_box.updated_at = datetime.now()

    db.session.commit()

    create_double_entry(
        '1001',          # الصندوق
        '4001',          # الإيرادات
        sale.total_amount,
        'SAL',
        sale.id,
        f'بيع نقدي من عميل {sale.customer.name}',
        current_user.id
    )

    flash(f'تمت الموافقة على عملية البيع النقدية رقم {sale.id} بمبلغ {sale.total_amount:,.2f} ريال', 'success')
    return redirect(url_for('sales'))

@app.route('/sales/delete/<int:id>')
@login_required
@permission_required(Permission.ADD_SALE)
def delete_sale(id):
    sale = SaleOrder.query.get_or_404(id)

    # التحقق: لا يمكن حذف عملية مكتملة أو موافق عليها
    if sale.cash_status == 'approved':
        flash('لا يمكن حذف عملية بيع تمت الموافقة عليها مسبقاً', 'danger')
        return redirect(url_for('sales'))

    if sale.status == 'completed':
        flash('لا يمكن حذف عملية بيع مكتملة', 'danger')
        return redirect(url_for('sales'))

    # حذف عناصر البيع أولاً
    for item in sale.items:
        db.session.delete(item)

    # حذف أمر البيع
    db.session.delete(sale)
    db.session.commit()

    flash('تم حذف عملية البيع بنجاح', 'success')
    return redirect(url_for('sales'))


@app.route('/sales/cancel/<int:id>')
@login_required
@permission_required(Permission.ADD_SALE)
def cancel_sale(id):
    """إلغاء عملية بيع (بدلاً من حذفها)"""
    sale = SaleOrder.query.get_or_404(id)

    # التحقق: لا يمكن إلغاء عملية مكتملة
    if sale.status == 'completed':
        flash('لا يمكن إلغاء عملية بيع مكتملة', 'danger')
        return redirect(url_for('sales'))

    # إذا كانت العملية قد تمت الموافقة عليها مسبقاً، نعيد المخزون
    if sale.cash_status == 'approved':
        for item in sale.items:
            product = Product.query.get(item.product_id)
            if product:
                product.quantity += item.quantity

    # تحديث الحالة إلى ملغاة
    sale.status = 'cancelled'
    sale.cash_status = 'rejected'
    sale.cash_rejection_reason = 'تم الإلغاء بواسطة المستخدم'

    db.session.commit()
    flash('تم إلغاء عملية البيع بنجاح', 'success')
    return redirect(url_for('sales'))


@app.route('/sales/print/<int:id>')
@login_required
@permission_required(Permission.VIEW_SALES)
def print_sale(id):
    """طباعة تفاصيل عملية البيع"""
    sale = SaleOrder.query.get_or_404(id)
    return render_template('sales/print.html', sale=sale)


# ==================== إدارة التحصيل ====================
from models import Permission  # تأكد من الاستيراد

@app.route('/collections')
@login_required
@permission_required(Permission.VIEW_COLLECTIONS)
def collections():
    # جلب جميع التحصيلات
    collections_list = Collection.query.order_by(Collection.collection_date.desc()).all()

    # جلب العملاء الذين عليهم مديونية (balance > 0)
    debt_customers = Customer.query.filter(Customer.balance > 0).all()

    # جلب المحصلين (collector + admin)
    collectors = User.query.filter(
        (User.role == UserRole.COLLECTOR.value) | (User.role == UserRole.ADMIN.value)
    ).all()

    # حساب تحصيلات اليوم (المعتمدة فقط)
    today = datetime.now().date()
    today_collections = Collection.query.filter(
        db.func.date(Collection.collection_date) == today,
        Collection.cash_status == 'approved'  # فقط المعتمدة
    ).all()
    today_collections_total = sum(c.amount for c in today_collections)

    # حساب إجمالي التحصيلات (المعتمدة فقط)
    approved_collections = [c for c in collections_list if c.cash_status == 'approved']
    total_collections = sum(c.amount for c in approved_collections)

    # حساب إجمالي المديونيات
    total_debts = sum(c.balance for c in debt_customers)

    # تحصيلات تنتظر الموافقة
    pending_collections = Collection.query.filter_by(cash_status='pending').count()

    # إجمالي العملاء
    total_customers = Customer.query.count()
    customers_with_debt = len(debt_customers)

    return render_template('collections/index.html',
                           collections=collections_list,
                           debt_customers=debt_customers,
                           collectors=collectors,
                           today_collections_total=today_collections_total,
                           total_collections=total_collections,
                           total_debts=total_debts,
                           pending_collections=pending_collections,
                           total_customers=total_customers,
                           Permission=Permission,  # ✅ إضافة هذا
                            customers_with_debt=customers_with_debt)
@app.route('/collections/add', methods=['POST'])
@login_required
@permission_required(Permission.ADD_COLLECTION)
def add_collection():
    customer_id = request.form.get('customer_id')
    sale_order_id = request.form.get('sale_order_id')
    amount = float(request.form.get('amount', 0))
    collector_id = request.form.get('collector_id')
    notes = request.form.get('notes', '')

    if not customer_id:
        flash('الرجاء اختيار عميل', 'danger')
        return redirect(url_for('collections'))

    if amount <= 0:
        flash('المبلغ يجب أن يكون أكبر من صفر', 'danger')
        return redirect(url_for('collections'))

    customer = Customer.query.get(customer_id)
    if not customer:
        flash('العميل غير موجود', 'danger')
        return redirect(url_for('collections'))

    if amount > customer.balance:
        flash(f'المبلغ المحصل ({amount:,.2f} ر.ي) يتجاوز المديونية ({customer.balance:,.2f} ر.ي)', 'danger')
        return redirect(url_for('collections'))

    if not collector_id:
        collector_id = current_user.id

    collection = Collection(
        customer_id=customer_id,
        sale_order_id=sale_order_id if sale_order_id else None,
        collector_id=collector_id,
        amount=amount,
        notes=notes,
        cash_status='pending'
    )
    db.session.add(collection)
    db.session.commit()

    flash(f'تم تسجيل تحصيل بمبلغ {amount:,.2f} ر.ي للعميل {customer.name} وتنتظر موافقة الصندوق', 'success')
    return redirect(url_for('collections'))

@app.route('/collections/customer/<int:id>/history')
@login_required
def customer_collections_history(id):
    """API لجلب تاريخ تحصيلات العميل"""
    customer = Customer.query.get_or_404(id)
    collections = Collection.query.filter_by(customer_id=id).order_by(Collection.collection_date.desc()).limit(10).all()

    return jsonify({
        'customer_name': customer.name,
        'balance': customer.balance,
        'collections': [{
            'date': c.collection_date.strftime('%Y-%m-%d %H:%M'),
            'amount': c.amount,
            'collector': c.collector.full_name if c.collector else '-',
            'status': 'معتمد' if c.cash_status == 'approved' else 'قيد المراجعة'
        } for c in collections]
    })


@app.route('/api/customer/<int:id>/open_invoices')
@login_required
def customer_open_invoices(id):
    """API لجلب الفواتير المفتوحة للعميل"""
    customer = Customer.query.get_or_404(id)

    open_invoices = SaleOrder.query.filter(
        SaleOrder.customer_id == id,
        SaleOrder.payment_type == 'credit',
        SaleOrder.total_amount > SaleOrder.paid_amount
    ).order_by(SaleOrder.sale_date.asc()).all()

    invoices = [{
        'id': inv.id,
        'date': inv.sale_date.strftime('%Y-%m-%d'),
        'total': inv.total_amount,
        'paid': inv.paid_amount,
        'remaining': inv.total_amount - inv.paid_amount
    } for inv in open_invoices]

    return jsonify({'invoices': invoices})
@app.route('/collections/approve/<int:id>', methods=['POST'])
@login_required
@permission_required(Permission.MANAGE_CASH)
def approve_collection(id):
    from accounting_utils import create_customer_collection_entry

    collection = Collection.query.get_or_404(id)

    if collection.cash_status == 'approved':
        flash('تم اعتماد التحصيل مسبقاً', 'warning')
        return redirect(url_for('collections'))

    customer = collection.customer

    try:
        collection.cash_status = 'approved'
        collection.cash_approved_by = current_user.id
        collection.cash_approved_at = datetime.now()

        # تحديث رصيد العميل في جدول customers
        customer.balance = float(customer.balance or 0) - float(collection.amount or 0)

        # تحديث الفاتورة المرتبطة إن وجدت
        if collection.sale_order_id:
            sale = SaleOrder.query.get(collection.sale_order_id)
            if sale:
                sale.paid_amount = float(sale.paid_amount or 0) + float(collection.amount or 0)
                # حساب المبلغ المتبقي بدون استخدام remaining_amount
                remaining = sale.total_amount - sale.paid_amount
                sale.status = 'completed' if remaining <= 0 else 'partial'
        else:
            # البحث عن فاتورة مفتوحة لنفس العميل
            unpaid_sale = SaleOrder.query.filter(
                SaleOrder.customer_id == customer.id,
                SaleOrder.payment_type == 'credit',
                SaleOrder.total_amount > SaleOrder.paid_amount
            ).order_by(SaleOrder.sale_date.asc()).first()
            if unpaid_sale:
                unpaid_sale.paid_amount = float(unpaid_sale.paid_amount or 0) + float(collection.amount or 0)
                remaining = unpaid_sale.total_amount - unpaid_sale.paid_amount
                unpaid_sale.status = 'completed' if remaining <= 0 else 'partial'

        db.session.flush()

        # إنشاء القيد باستخدام حساب العميل الفرعي
        ok = create_customer_collection_entry(
            customer_id=customer.id,
            amount=float(collection.amount or 0),
            reference_type='COL',
            reference_id=collection.id,
            description=f'تحصيل من العميل {customer.name}',
            user_id=current_user.id
        )

        if not ok:
            raise Exception("فشل إنشاء القيد المحاسبي")

        # تسجيل حركة الصندوق
        cash_transaction = CashTransaction(
            type='income',
            amount=collection.amount,
            description=f'تحصيل من العميل {customer.name}',
            reference_type='collection',
            reference_id=collection.id,
            user_id=current_user.id
        )
        db.session.add(cash_transaction)

        # تحديث رصيد الصندوق
        cash_box = CashBox.query.first()
        if cash_box:
            cash_box.balance += collection.amount
            cash_box.updated_at = datetime.now()

        db.session.commit()
        flash(f'تم اعتماد التحصيل بمبلغ {collection.amount:,.2f} ريال', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء اعتماد التحصيل: {str(e)}', 'danger')

    return redirect(url_for('collections'))

@app.route('/collections/reject/<int:id>', methods=['POST'])
@login_required
@permission_required(Permission.MANAGE_CASH)
def reject_collection(id):
    """رفض تحصيل"""
    collection = Collection.query.get_or_404(id)

    if collection.cash_status == 'approved':
        flash('لا يمكن رفض تحصيل تم اعتماده مسبقاً', 'danger')
        return redirect(url_for('collections'))

    collection.cash_status = 'rejected'
    collection.cash_rejection_reason = request.form.get('reason', '')

    db.session.commit()

    flash('تم رفض التحصيل', 'warning')
    return redirect(url_for('collections'))

# ==================== أمانات التجميد ====================

@app.route('/freeze-deposits')
@login_required
def freeze_deposits_list():
    deposits = FreezeDeposit.query.order_by(FreezeDeposit.deposit_date.desc()).all()
    return render_template('freeze_deposits/index.html', deposits=deposits)


@app.route('/freeze-deposits/add', methods=['GET', 'POST'])
@login_required
def add_freeze_deposit():
    if request.method == 'POST':
        product_name = request.form.get('product_name')
        category = request.form.get('category')
        quantity = int(request.form.get('quantity'))
        freeze_deposit_amount = float(request.form.get('freeze_deposit_amount'))
        selling_price = float(request.form.get('selling_price', 0))
        customer_id = request.form.get('customer_id')
        notes = request.form.get('notes', '')

        existing_product = Product.query.filter_by(name=product_name).first()

        if existing_product:
            existing_product.quantity += quantity
            product = existing_product
            flash(f'تم إضافة {quantity} وحدة للمنتج {product_name} كأمانة تجميد', 'info')
        else:
            product = Product(
                name=product_name,
                category=category,
                unit=0,
                purchase_price=0,
                selling_price=selling_price,
                quantity=quantity,
                min_quantity=10,
                location='ثلاجة التجميد',
                is_frozen=True,
                freeze_deposit=freeze_deposit_amount
            )
            db.session.add(product)
            db.session.flush()
            flash(f'تم إنشاء المنتج {product_name} مع أمانة تجميد', 'success')

        deposit = FreezeDeposit(
            product_id=product.id,
            customer_id=customer_id if customer_id else None,
            amount=freeze_deposit_amount * quantity,
            quantity=quantity,
            notes=notes,
            created_by=current_user.id
        )
        db.session.add(deposit)

        cash_transaction = CashTransaction(
            type='deposit',
            amount=freeze_deposit_amount * quantity,
            description=f'أمانة تجميد للمنتج {product_name} (كمية: {quantity})',
            reference_type='freeze_deposit',
            reference_id=deposit.id,
            user_id=current_user.id
        )
        db.session.add(cash_transaction)

        db.session.commit()
        return redirect(url_for('freeze_deposits_list'))

    customers = Customer.query.all()
    categories = ['water', 'soda']
    return render_template('freeze_deposits/add.html', customers=customers, categories=categories)


@app.route('/freeze-deposits/return/<int:id>', methods=['POST'])
@login_required
def return_freeze_deposit(id):
    deposit = FreezeDeposit.query.get_or_404(id)
    if deposit.is_returned:
        flash('هذه الأمانة تم إرجاعها بالفعل', 'warning')
        return redirect(url_for('freeze_deposits_list'))

    deposit.is_returned = True
    deposit.return_date = datetime.now()

    cash_transaction = CashTransaction(
        type='deposit_return',
        amount=deposit.amount,
        description=f'استرداد أمانة تجميد للمنتج {deposit.product.name}',
        reference_type='freeze_deposit_return',
        reference_id=deposit.id,
        user_id=current_user.id
    )
    db.session.add(cash_transaction)

    db.session.commit()
    flash(f'تم إرجاع أمانة التجميد بمبلغ {deposit.amount} د.ع', 'success')
    return redirect(url_for('freeze_deposits_list'))


# ==================== إدارة الصندوق ====================
@app.route('/cash')
@login_required
@permission_required(Permission.VIEW_CASH)
def cash_index():
    from sqlalchemy import func
    cash_box = CashBox.query.first()
    if not cash_box:
        cash_box = CashBox(name='الصندوق الرئيسي', balance=0, initial_balance=0)
        db.session.add(cash_box)
        db.session.commit()

    today = datetime.now().date()

    # المبيعات النقدية المعتمدة اليوم
    today_cash_sales = db.session.query(func.sum(SaleOrder.total_amount)).filter(
        func.date(SaleOrder.sale_date) == today,
        SaleOrder.payment_type == 'cash',
        SaleOrder.cash_status == 'approved'
    ).scalar() or 0

    # المبيعات الآجلة المعتمدة اليوم
    today_credit_sales = db.session.query(func.sum(SaleOrder.total_amount)).filter(
        func.date(SaleOrder.sale_date) == today,
        SaleOrder.payment_type == 'credit',
        SaleOrder.cash_status == 'approved'
    ).scalar() or 0

    # التحصيلات المعتمدة اليوم
    today_collections = db.session.query(func.sum(Collection.amount)).filter(
        func.date(Collection.collection_date) == today,
        Collection.cash_status == 'approved'
    ).scalar() or 0

    # المشتريات النقدية المعتمدة اليوم
    today_purchases = db.session.query(func.sum(PurchaseOrder.total_amount)).filter(
        func.date(PurchaseOrder.order_date) == today,
        PurchaseOrder.payment_type == 'cash',
        PurchaseOrder.cash_status == 'approved'
    ).scalar() or 0

    # أمانات التجميد المضافة اليوم
    today_deposits = db.session.query(func.sum(FreezeDeposit.amount)).filter(
        func.date(FreezeDeposit.deposit_date) == today
    ).scalar() or 0

    # مدفوعات الموردين المعتمدة اليوم
    today_supplier_payments = db.session.query(func.sum(SupplierPayment.amount)).filter(
        func.date(SupplierPayment.payment_date) == today,
        SupplierPayment.cash_status == 'approved'
    ).scalar() or 0

    # حساب إحصائيات الشهر الحالي
    first_day_of_month = datetime.now().replace(day=1)

    # المبيعات النقدية المعتمدة هذا الشهر
    monthly_cash_sales = db.session.query(func.sum(SaleOrder.total_amount)).filter(
        SaleOrder.sale_date >= first_day_of_month,
        SaleOrder.payment_type == 'cash',
        SaleOrder.cash_status == 'approved'
    ).scalar() or 0

    # التحصيلات المعتمدة هذا الشهر
    monthly_collections = db.session.query(func.sum(Collection.amount)).filter(
        Collection.collection_date >= first_day_of_month,
        Collection.cash_status == 'approved'
    ).scalar() or 0

    # المشتريات النقدية المعتمدة هذا الشهر
    monthly_purchases = db.session.query(func.sum(PurchaseOrder.total_amount)).filter(
        PurchaseOrder.order_date >= first_day_of_month,
        PurchaseOrder.payment_type == 'cash',
        PurchaseOrder.cash_status == 'approved'
    ).scalar() or 0

    # مدفوعات الموردين المعتمدة هذا الشهر
    monthly_supplier_payments = db.session.query(func.sum(SupplierPayment.amount)).filter(
        SupplierPayment.payment_date >= first_day_of_month,
        SupplierPayment.cash_status == 'approved'
    ).scalar() or 0

    # إجمالي الإيرادات هذا الشهر
    monthly_total_income = monthly_cash_sales + monthly_collections

    # إجمالي المصروفات هذا الشهر
    monthly_total_expense = monthly_purchases + monthly_supplier_payments

    # صافي هذا الشهر
    monthly_net = monthly_total_income - monthly_total_expense

    total_income = today_cash_sales + today_collections + today_deposits
    total_expense = today_purchases + today_supplier_payments
    total_net = total_income - total_expense

    # تحديث رصيد الصندوق الفعلي من المعاملات
    total_income_all = db.session.query(func.sum(CashTransaction.amount)).filter(CashTransaction.type == 'income').scalar() or 0
    total_expense_all = db.session.query(func.sum(CashTransaction.amount)).filter(CashTransaction.type == 'expense').scalar() or 0
    calculated_balance = total_income_all - total_expense_all
    if abs(cash_box.balance - calculated_balance) > 0.01:
        cash_box.balance = calculated_balance
        cash_box.updated_at = datetime.now()
        db.session.commit()

    today_summary = DailyCashSummary.query.filter(func.date(DailyCashSummary.summary_date) == today).first()
    pending_count = PurchaseOrder.query.filter_by(cash_status='pending', payment_type='cash').count()
    pending_count += SaleOrder.query.filter_by(cash_status='pending', payment_type='cash').count()
    pending_count += Collection.query.filter_by(cash_status='pending').count()

    return render_template('cash/index.html',
                           cash_box=cash_box,
                           today_summary=today_summary,
                           today_cash_sales=today_cash_sales,
                           today_credit_sales=today_credit_sales,
                           today_collections=today_collections,
                           today_purchases=today_purchases,
                           today_deposits=today_deposits,
                           today_supplier_payments=today_supplier_payments,
                           total_income=total_income,
                           total_expense=total_expense,
                           total_net=total_net,
                           pending_count=pending_count,
                           monthly_cash_sales=monthly_cash_sales,
                           monthly_collections=monthly_collections,
                           monthly_purchases=monthly_purchases,
                           monthly_supplier_payments=monthly_supplier_payments,
                           monthly_total_income=monthly_total_income,
                           monthly_total_expense=monthly_total_expense,
                           monthly_net=monthly_net)

from io import BytesIO
from datetime import datetime

import pandas as pd
from flask import render_template, request, make_response, flash, redirect, url_for
from sqlalchemy import or_
from weasyprint import HTML

# إذا كان عندك استيرادات إضافية موجودة مسبقًا اتركها كما هي
# من الأفضل أن يكون JournalEntry و JournalDetail معرفة عندك بالفعل

from sqlalchemy import or_
from sqlalchemy import or_

from sqlalchemy import or_

from sqlalchemy import or_

from sqlalchemy import or_, exists
from datetime import datetime

def build_journal_entries_query():
    query = JournalEntry.query

    from_date = request.args.get('from_date', '').strip()
    to_date = request.args.get('to_date', '').strip()
    account_name = request.args.get('account_name', '').strip()
    account_code = request.args.get('account_code', '').strip()

    if from_date:
        try:
            start_dt = datetime.strptime(from_date, '%Y-%m-%d')
            query = query.filter(JournalEntry.entry_date >= start_dt)
        except ValueError:
            flash('صيغة "من تاريخ" غير صحيحة', 'warning')

    if to_date:
        try:
            end_dt = datetime.strptime(to_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            query = query.filter(JournalEntry.entry_date <= end_dt)
        except ValueError:
            flash('صيغة "إلى تاريخ" غير صحيحة', 'warning')

    # إذا تم إدخال اسم حساب أو رقم حساب
    if account_name or account_code:
        search_term = account_name or account_code
        filters = []

        # 1. البحث في وصف القيد (البيان)
        filters.append(JournalEntry.description.ilike(f'%{search_term}%'))

        # 2. البحث في اسم الحساب أو رقم الحساب عبر العلاقة مع financial_accounts
        account_exists = exists().where(
            JournalDetail.entry_id == JournalEntry.id,
            JournalDetail.account_id == FinancialAccount.id,
            or_(
                FinancialAccount.account_name.ilike(f'%{search_term}%'),
                FinancialAccount.account_code.ilike(f'%{search_term}%')
            )
        )
        filters.append(account_exists)

        # تطبيق الفلاتر ومنع التكرار
        query = query.filter(or_(*filters)).distinct(JournalEntry.id)

    return query

@app.route('/cash/journal')
@login_required
@permission_required(Permission.VIEW_JOURNAL)
def journal_entries():
    page = request.args.get('page', 1, type=int)
    query = build_journal_entries_query()

    # تأكد من أن النتائج مميزة (لا تكرار) قبل الترقيم
    query = query.order_by(JournalEntry.entry_date.desc()).distinct(JournalEntry.id)

    entries = query.paginate(page=page, per_page=15, error_out=False)

    return render_template('cash/journal.html', entries=entries)

@app.route('/cash/journal/add', methods=['GET', 'POST'])
@login_required
@permission_required(Permission.ADD_JOURNAL_ENTRY)
def add_journal_entry():
    """إضافة قيد يومية جديد"""
    if request.method == 'POST':
        reference_number = request.form.get('reference_number')
        description = request.form.get('description')

        if not reference_number:
            reference_number = f"ENT-{datetime.now().strftime('%Y%m%d%H%M%S')}"

        entry = JournalEntry(
            reference_number=reference_number,
            description=description,
            created_by=current_user.id
        )
        db.session.add(entry)
        db.session.flush()

        total_debit = 0
        total_credit = 0

        accounts = request.form.getlist('account_type[]')
        account_names = request.form.getlist('account_name[]')
        amounts = request.form.getlist('amount[]')
        types = request.form.getlist('dc_type[]')
        notes_list = request.form.getlist('notes[]')

        for i in range(len(accounts)):
            if i < len(accounts) and i < len(account_names) and i < len(amounts) and accounts[i] and amounts[i]:
                amount = float(amounts[i])

                detail = JournalDetail(
                    entry_id=entry.id,
                    account_type=accounts[i],
                    account_name=account_names[i],
                    debit=amount if types[i] == 'debit' else 0,
                    credit=amount if types[i] == 'credit' else 0,
                    notes=notes_list[i] if i < len(notes_list) else ''
                )
                db.session.add(detail)

                if types[i] == 'debit':
                    total_debit += amount
                else:
                    total_credit += amount

        if total_debit != total_credit:
            flash('القيد غير متوازن! يجب أن يكون مجموع المدين = مجموع الدائن', 'danger')
            db.session.rollback()
            return redirect(url_for('add_journal_entry'))

        entry.total_debit = total_debit
        entry.total_credit = total_credit

        cash_box = CashBox.query.first()
        if cash_box:
            for detail in entry.details:
                if detail.account_type == 'cash':
                    if detail.debit > 0:
                        cash_box.balance += detail.debit
                    elif detail.credit > 0:
                        cash_box.balance -= detail.credit
            cash_box.updated_at = datetime.now()

        db.session.commit()
        flash('تم إضافة قيد اليومية بنجاح', 'success')
        return redirect(url_for('journal_entries'))

    return render_template('cash/add_journal.html')


@app.route('/cash/journal/delete/<int:id>')
@login_required
@permission_required(Permission.MANAGE_CASH)
def delete_journal_entry(id):
    """حذف قيد يومية"""
    entry = JournalEntry.query.get_or_404(id)

    cash_box = CashBox.query.first()
    if cash_box:
        for detail in entry.details:
            if detail.account_type == 'cash':
                if detail.debit > 0:
                    cash_box.balance -= detail.debit
                elif detail.credit > 0:
                    cash_box.balance += detail.credit

    db.session.delete(entry)
    db.session.commit()
    flash('تم حذف قيد اليومية بنجاح', 'success')
    return redirect(url_for('journal_entries'))



@app.route('/journal/export/excel')
@login_required
@permission_required(Permission.VIEW_JOURNAL)
def export_journal_excel():
    """تصدير Excel بنفس الفلاتر"""
    query = build_journal_entries_query()
    entries = query.order_by(JournalEntry.entry_date.desc()).all()

    rows = []
    for entry in entries:
        rows.append({
            'رقم القيد': entry.id,
            'التاريخ': entry.entry_date.strftime('%Y-%m-%d %H:%M') if entry.entry_date else '',
            'رقم الإذن': entry.reference_number or '',
            'البيان': entry.description or '',
            'إجمالي المدين': float(entry.total_debit or 0),
            'إجمالي الدائن': float(entry.total_credit or 0),
            'أنشئ بواسطة': entry.creator.full_name if getattr(entry, 'creator', None) else 'النظام',
        })

    df = pd.DataFrame(rows)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='JournalEntries')

    output.seek(0)

    response = make_response(output.read())
    response.headers['Content-Disposition'] = 'attachment; filename=journal_entries.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/journal/export/pdf')
@login_required
@permission_required(Permission.VIEW_JOURNAL)
def export_journal_pdf():
    """تصدير PDF بنفس الفلاتر"""
    query = build_journal_entries_query()
    entries = query.order_by(JournalEntry.entry_date.desc()).all()

    html = render_template(
        'cash/journal_entries_pdf.html',
        entries=entries,
        from_date=request.args.get('from_date', ''),
        to_date=request.args.get('to_date', ''),
        account_name=request.args.get('account_name', ''),
        account_code=request.args.get('account_code', '')
    )

    pdf_bytes = HTML(
        string=html,
        base_url=request.url_root
    ).write_pdf()

    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=journal_entries.pdf'
    return response

@app.route('/cash/opening-balance', methods=['POST'])
@login_required
@permission_required(Permission.MANAGE_CASH)
def set_opening_balance():
    """تحديد الرصيد الافتتاحي للصندوق"""
    amount = float(request.form.get('opening_balance', 0))
    cash_box = CashBox.query.first()
    if cash_box:
        cash_box.initial_balance = amount
        cash_box.balance = amount
        cash_box.updated_at = datetime.now()
        db.session.commit()
        flash(f'تم تحديد الرصيد الافتتاحي للصندوق بمبلغ {amount} د.ع', 'success')
    return redirect(url_for('cash_index'))


@app.route('/cash/daily-closing', methods=['POST'])
@login_required
@permission_required(Permission.DAILY_CLOSING)
def daily_closing():
    """إغلاق يومي للصندوق"""
    today = datetime.now().date()

    # التحقق من عدم وجود إغلاق سابق لنفس اليوم
    existing = DailyCashSummary.query.filter(db.func.date(DailyCashSummary.summary_date) == today).first()
    if existing:
        flash('تم إغلاق الصندوق لهذا اليوم مسبقاً', 'warning')
        return redirect(url_for('cash_index'))

    cash_box = CashBox.query.first()

    # حساب إحصائيات اليوم
    today_cash_sales = db.session.query(db.func.sum(SaleOrder.paid_amount)).filter(
        db.func.date(SaleOrder.sale_date) == today,
        SaleOrder.payment_type == 'cash'
    ).scalar() or 0

    today_credit_sales = db.session.query(db.func.sum(SaleOrder.total_amount)).filter(
        db.func.date(SaleOrder.sale_date) == today,
        SaleOrder.payment_type == 'credit'
    ).scalar() or 0

    today_collections = db.session.query(db.func.sum(Collection.amount)).filter(
        db.func.date(Collection.collection_date) == today
    ).scalar() or 0

    today_purchases = db.session.query(db.func.sum(PurchaseOrder.paid_amount)).filter(
        db.func.date(PurchaseOrder.order_date) == today
    ).scalar() or 0

    today_deposits = db.session.query(db.func.sum(CashTransaction.amount)).filter(
        db.func.date(CashTransaction.transaction_date) == today,
        CashTransaction.type == 'deposit'
    ).scalar() or 0

    total_income = today_cash_sales + today_collections + today_deposits
    total_expense = today_purchases
    closing_balance = cash_box.balance if cash_box else 0

    summary = DailyCashSummary(
        summary_date=datetime.now(),
        opening_balance=cash_box.initial_balance if cash_box else 0,
        closing_balance=closing_balance,
        total_income=total_income,
        total_expense=total_expense,
        cash_sales=today_cash_sales,
        credit_sales=today_credit_sales,
        collections=today_collections,
        purchases=today_purchases,
        deposits=today_deposits,
        created_by=current_user.id
    )
    db.session.add(summary)
    db.session.commit()

    flash('تم إغلاق الصندوق اليومي بنجاح', 'success')
    return redirect(url_for('cash_index'))


@app.route('/cash/transactions')
@login_required
@permission_required(Permission.VIEW_CASH)
def cash_transactions():
    """سجل معاملات الصندوق"""
    transactions = CashTransaction.query.order_by(CashTransaction.transaction_date.desc()).all()
    return render_template('cash/transactions.html', transactions=transactions)


@app.route('/cash/approvals')
@login_required
@permission_required(Permission.MANAGE_CASH)
def cash_approvals():
    """عرض المعاملات التي تنتظر موافقة الصندوق"""

    # ✅ مشتريات نقدية كاملة
    cash_purchases = PurchaseOrder.query.filter_by(
        cash_status='pending',
        payment_type='cash'
    ).all()

    # ✅ مشتريات آجلة مع دفعة مقدمة
    advance_purchases = PurchaseOrder.query.filter(
        PurchaseOrder.cash_status == 'pending',
        PurchaseOrder.payment_type == 'credit',
        PurchaseOrder.paid_amount == 0
    ).all()

    # ✅ مبيعات نقدية
    pending_sales = SaleOrder.query.filter_by(
        cash_status='pending',
        payment_type='cash'
    ).all()

    # ✅ تحصيلات
    pending_collections = Collection.query.filter_by(cash_status='pending').all()

    # ✅ دمج المشتريات في قائمة واحدة للعرض
    pending_purchases = cash_purchases + advance_purchases

    return render_template('cash/approvals.html',
                           pending_purchases=pending_purchases,
                           pending_sales=pending_sales,
                           pending_collections=pending_collections)

@app.route('/cash/approve/<string:type>/<int:id>', methods=['POST'])
@login_required
@permission_required(Permission.MANAGE_CASH)
def approve_transaction(type, id):
    """الموافقة على معاملة مالية - للمعاملات النقدية والدفعات المقدمة"""
    from accounting_utils import create_double_entry

    action = request.form.get('action')
    reason = request.form.get('reason', '')
    cash_amount = float(request.form.get('cash_amount', 0))

    if type == 'purchase':
        transaction = PurchaseOrder.query.get_or_404(id)
        if transaction.payment_type == 'cash':
            cash_amount = transaction.total_amount
        elif transaction.payment_type == 'credit' and transaction.paid_amount == 0 and cash_amount > 0:
            pass
        else:
            flash('هذه المعاملة لا تحتاج موافقة الصندوق', 'warning')
            return redirect(url_for('cash_approvals'))

    elif type == 'sale':
        transaction = SaleOrder.query.get_or_404(id)
        if transaction.payment_type != 'cash':
            flash('المعاملات الآجلة لا تحتاج موافقة الصندوق', 'warning')
            return redirect(url_for('cash_approvals'))
        cash_amount = transaction.total_amount

    elif type == 'collection':
        transaction = Collection.query.get_or_404(id)
        cash_amount = transaction.amount
    else:
        flash('نوع المعاملة غير صحيح', 'danger')
        return redirect(url_for('cash_approvals'))

    if action == 'approve':
        transaction.cash_status = 'approved'
        transaction.cash_approved_by = current_user.id
        transaction.cash_approved_at = datetime.now()

        if type == 'purchase':
            # تحديث المخزون
            for item in transaction.items:
                product = Product.query.get(item.product_id)
                if product:
                    product.quantity += item.quantity
                    product.purchase_price = item.unit_price

            # تحديث المبلغ المدفوع
            if transaction.payment_type == 'cash':
                transaction.paid_amount = transaction.total_amount
            else:
                transaction.paid_amount = cash_amount

            # تحديث رصيد المورد
            if transaction.remaining_amount > 0:
                supplier = Supplier.query.get(transaction.supplier_id)
                if supplier:
                    supplier.balance += transaction.remaining_amount

            db.session.commit()

            # القيد المحاسبي للشراء
            if cash_amount > 0:
                if transaction.payment_type == 'cash':
                    create_double_entry('5001', '1001', cash_amount, 'PUR', transaction.id,
                                        f'شراء نقدي من مورد {transaction.supplier.name}', current_user.id)
                else:
                    create_double_entry('5001', '1001', cash_amount, 'PUR', transaction.id,
                                        f'دفعة مقدمة لمورد {transaction.supplier.name}', current_user.id)

            transaction.status = 'completed'
            flash(f'تمت الموافقة على عملية الشراء بقيمة {cash_amount:,.2f} ريال', 'success')

        elif type == 'sale':
            # تحديث المخزون
            for item in transaction.items:
                product = Product.query.get(item.product_id)
                if product:
                    product.quantity -= item.quantity

            transaction.paid_amount = transaction.total_amount
            transaction.remaining_amount = 0

            db.session.commit()

            # القيد المحاسبي للبيع النقدي
            create_double_entry('1001', '4001', cash_amount, 'SAL', transaction.id,
                                f'بيع نقدي من عميل {transaction.customer.name}', current_user.id)

            # تسجيل حركة الصندوق
            cash_transaction = CashTransaction(
                type='income',
                amount=cash_amount,
                description=f'بيع نقدي من عميل {transaction.customer.name}',
                reference_type='sale',
                reference_id=transaction.id,
                user_id=current_user.id
            )
            db.session.add(cash_transaction)

            # تحديث رصيد الصندوق
            cash_box = CashBox.query.first()
            if cash_box:
                cash_box.balance += cash_amount
                cash_box.updated_at = datetime.now()

            transaction.status = 'completed'
            flash(f'تمت الموافقة على عملية البيع النقدية بقيمة {cash_amount:,.2f} ريال', 'success')

        elif type == 'collection':
            customer = Customer.query.get(transaction.customer_id)
            if customer:
                # خصم المبلغ من رصيد العميل
                customer.balance -= cash_amount
                db.session.commit()

            # ✅ تحديث الفاتورة المرتبطة مباشرة (بدلاً من البحث)
            if transaction.sale_order_id:
                sale = SaleOrder.query.get(transaction.sale_order_id)
                if sale:
                    sale.paid_amount = (sale.paid_amount or 0) + cash_amount
                    sale.remaining_amount = sale.total_amount - sale.paid_amount
                    if sale.paid_amount >= sale.total_amount:
                        sale.status = 'completed'
                    else:
                        sale.status = 'partial'
                    db.session.commit()
                    print(f"✅ تم تحديث فاتورة {sale.id}: مدفوع جديد {sale.paid_amount:,.2f} ريال")
            else:
                # البحث عن فاتورة مفتوحة لنفس العميل
                unpaid_sale = SaleOrder.query.filter(
                    SaleOrder.customer_id == transaction.customer_id,
                    SaleOrder.payment_type == 'credit',
                    SaleOrder.total_amount > SaleOrder.paid_amount,
                    SaleOrder.status == 'completed'
                ).order_by(SaleOrder.sale_date.asc()).first()

                if unpaid_sale:
                    unpaid_sale.paid_amount = (unpaid_sale.paid_amount or 0) + cash_amount
                    if unpaid_sale.paid_amount >= unpaid_sale.total_amount:
                        unpaid_sale.paid_amount = unpaid_sale.total_amount
                        unpaid_sale.remaining_amount = 0
                        unpaid_sale.status = 'completed'
                    else:
                        unpaid_sale.remaining_amount = unpaid_sale.total_amount - unpaid_sale.paid_amount
                        unpaid_sale.status = 'partial'
                    db.session.commit()
                    print(f"✅ تم تحديث فاتورة {unpaid_sale.id}: مدفوع جديد {unpaid_sale.paid_amount:,.2f} ريال")

            # ✅ القيد المحاسبي للتحصيل
            create_customer_collection_entry(
                customer.id,
                cash_amount,
                'COL',
                transaction.id,
                f'تحصيل من عميل {customer.name}',
                current_user.id
            )

            # تسجيل حركة إيداع في الصندوق
            cash_transaction = CashTransaction(
                type='income',
                amount=cash_amount,
                description=f'تحصيل من عميل {customer.name}',
                reference_type='collection',
                reference_id=transaction.id,
                user_id=current_user.id
            )
            db.session.add(cash_transaction)

            # تحديث رصيد الصندوق
            cash_box = CashBox.query.first()
            if cash_box:
                cash_box.balance += cash_amount
                cash_box.updated_at = datetime.now()

            transaction.status = 'completed'
            flash(f'تمت الموافقة على عملية التحصيل بقيمة {cash_amount:,.2f} ريال', 'success')

    else:  # reject
        transaction.cash_status = 'rejected'
        transaction.cash_rejection_reason = reason
        transaction.status = 'cancelled'

        if type == 'purchase' and transaction.payment_type == 'credit':
            for item in transaction.items:
                db.session.delete(item)

        flash(f'تم رفض المعاملة: {reason}', 'warning')

    db.session.commit()
    return redirect(url_for('cash_approvals'))

@app.route('/reports')
@login_required
@permission_required(Permission.VIEW_REPORTS)
def reports():
    return render_template('reports/index.html')


@app.route('/api/products/search')
@login_required
def search_products():
    query = request.args.get('q', '')
    products = Product.query.filter(Product.name.contains(query)).limit(10).all()
    return jsonify([{'id': p.id, 'name': p.name, 'price': p.selling_price, 'quantity': p.quantity} for p in products])


@app.route('/api/reports/data')
@login_required
@permission_required(Permission.VIEW_REPORTS)
def reports_data():
    """API لجلب بيانات التقارير"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    report_type = request.args.get('type', 'all')

    # تحويل التواريخ
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
    else:
        start_date = datetime.now().replace(day=1)

    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
    else:
        end_date = datetime.now()

    # جلب بيانات المبيعات
    sales_query = SaleOrder.query.filter(
        db.func.date(SaleOrder.sale_date) >= start_date.date(),
        db.func.date(SaleOrder.sale_date) <= end_date.date()
    )
    sales = sales_query.all()

    sales_data = [{
        'date': s.sale_date.strftime('%Y-%m-%d'),
        'customer': s.customer.name,
        'total': s.total_amount,
        'paid': s.paid_amount,
        'remaining': s.total_amount - s.paid_amount,
        'type': s.payment_type
    } for s in sales]

    total_sales = sum(s.total_amount for s in sales)
    total_collected = sum(s.paid_amount for s in sales)

    # جلب بيانات المشتريات
    purchases_query = PurchaseOrder.query.filter(
        db.func.date(PurchaseOrder.order_date) >= start_date.date(),
        db.func.date(PurchaseOrder.order_date) <= end_date.date()
    )
    purchases = purchases_query.all()

    purchases_data = [{
        'date': p.order_date.strftime('%Y-%m-%d'),
        'supplier': p.supplier.name,
        'total': p.total_amount,
        'paid': p.paid_amount,
        'remaining': p.total_amount - p.paid_amount,
        'type': p.payment_type
    } for p in purchases]

    total_purchases = sum(p.total_amount for p in purchases)

    # جلب بيانات المديونيات
    debt_customers = Customer.query.filter(Customer.balance > 0).all()
    debts_data = [{
        'customer': c.name,
        'balance': c.balance,
        'limit': c.credit_limit,
        'percentage': (c.balance / c.credit_limit * 100) if c.credit_limit > 0 else 0,
        'last_collection': None
    } for c in debt_customers]

    total_debts = sum(c.balance for c in debt_customers)

    # جلب بيانات الصندوق
    cash_transactions = CashTransaction.query.filter(
        db.func.date(CashTransaction.transaction_date) >= start_date.date(),
        db.func.date(CashTransaction.transaction_date) <= end_date.date()
    ).all()

    cash_data = [{
        'date': t.transaction_date.strftime('%Y-%m-%d %H:%M'),
        'type': t.type,
        'amount': t.amount,
        'description': t.description,
        'user': t.user.full_name if t.user else '-'
    } for t in cash_transactions]

    total_expenses = sum(t.amount for t in cash_transactions if t.type == 'expense')
    total_collections = sum(t.amount for t in cash_transactions if t.type == 'income')

    # جلب بيانات المخزون
    products = Product.query.all()
    inventory_data = [{
        'name': p.name,
        'quantity': p.quantity,
        'unit': p.unit,
        'purchase_price': p.purchase_price,
        'selling_price': p.selling_price,
        'value': p.quantity * p.purchase_price,
        'min_quantity': p.min_quantity
    } for p in products]

    # جلب بيانات أمانات التجميد
    freeze_deposits = FreezeDeposit.query.filter(
        db.func.date(FreezeDeposit.deposit_date) >= start_date.date(),
        db.func.date(FreezeDeposit.deposit_date) <= end_date.date()
    ).all()

    freeze_data = [{
        'date': d.deposit_date.strftime('%Y-%m-%d'),
        'product': d.product.name,
        'quantity': d.quantity,
        'amount': d.amount,
        'is_returned': d.is_returned,
        'customer': d.customer.name if d.customer else '-'
    } for d in freeze_deposits]

    # البيانات الشهرية للرسوم البيانية
    months = ['يناير', 'فبراير', 'مارس', 'إبريل', 'مايو', 'يونيو', 'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر',
              'ديسمبر']
    monthly_sales = []
    monthly_purchases = []

    for i in range(1, 13):
        month_sales = SaleOrder.query.filter(
            db.extract('month', SaleOrder.sale_date) == i,
            db.extract('year', SaleOrder.sale_date) == datetime.now().year
        ).all()
        monthly_sales.append(sum(s.total_amount for s in month_sales))

        month_purchases = PurchaseOrder.query.filter(
            db.extract('month', PurchaseOrder.order_date) == i,
            db.extract('year', PurchaseOrder.order_date) == datetime.now().year
        ).all()
        monthly_purchases.append(sum(p.total_amount for p in month_purchases))

    return jsonify({
        'total_sales': total_sales,
        'total_purchases': total_purchases,
        'total_profit': total_sales - total_purchases,
        'total_debts': total_debts,
        'total_expenses': total_expenses,
        'total_collections': total_collections,
        'sales': sales_data,
        'purchases': purchases_data,
        'debts': debts_data,
        'cash_transactions': cash_data,
        'inventory': inventory_data,
        'freeze_deposits': freeze_data,
        'months': months[:6],
        'monthly_sales': monthly_sales[:6],
        'monthly_purchases': monthly_purchases[:6]
    })


@app.route('/api/customer/<int:id>/details')
@login_required
def customer_details(id):
    """جلب تفاصيل العميل للمودال"""
    customer = Customer.query.get_or_404(id)

    # جلب عمليات الشراء للعميل
    sales = SaleOrder.query.filter_by(customer_id=id).order_by(SaleOrder.sale_date.desc()).limit(10).all()
    collections = Collection.query.filter_by(customer_id=id).order_by(Collection.collection_date.desc()).limit(10).all()

    transactions = []
    for sale in sales:
        transactions.append({
            'date': sale.sale_date.strftime('%Y-%m-%d'),
            'type': 'sale',
            'amount': sale.total_amount,
            'description': f'فاتورة رقم {sale.id}'
        })

    for coll in collections:
        transactions.append({
            'date': coll.collection_date.strftime('%Y-%m-%d'),
            'type': 'collection',
            'amount': coll.amount,
            'description': coll.notes or 'تحصيل نقدي'
        })

    # ترتيب المعاملات حسب التاريخ
    transactions.sort(key=lambda x: x['date'], reverse=True)

    total_purchases = sum(s.total_amount for s in sales)
    total_payments = sum(c.amount for c in collections)

    return jsonify({
        'id': customer.id,
        'name': customer.name,
        'phone': customer.phone,
        'address': customer.address,
        'credit_limit': customer.credit_limit,
        'balance': customer.balance,
        'total_purchases': total_purchases,
        'total_payments': total_payments,
        'last_purchase': sales[0].sale_date.strftime('%Y-%m-%d') if sales else None,
        'last_collection': collections[0].collection_date.strftime('%Y-%m-%d') if collections else None,
        'transactions': transactions[:20]
    })


@app.route('/reports/sales')
@login_required
@permission_required(Permission.VIEW_REPORTS)
def sales_report():
    start_date = request.args.get('start_date', datetime.now().strftime('%Y-%m-01'))
    end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
    payment_type = request.args.get('payment_type', 'all')

    query = SaleOrder.query.filter(
        db.func.date(SaleOrder.sale_date) >= start_date,
        db.func.date(SaleOrder.sale_date) <= end_date
    )

    if payment_type != 'all':
        query = query.filter(SaleOrder.payment_type == payment_type)

    sales_data = query.order_by(SaleOrder.sale_date.desc()).all()

    total_sales = sum(s.total_amount for s in sales_data)
    total_collected = sum(s.paid_amount for s in sales_data)
    total_credit = total_sales - total_collected

    # بيانات الرسم البياني اليومي
    daily_data = db.session.query(
        db.func.date(SaleOrder.sale_date).label('date'),
        db.func.sum(SaleOrder.total_amount).label('total')
    ).filter(
        db.func.date(SaleOrder.sale_date) >= start_date,
        db.func.date(SaleOrder.sale_date) <= end_date
    ).group_by(db.func.date(SaleOrder.sale_date)).all()

    daily_sales = [{'date': d.date.strftime('%Y-%m-%d'), 'total': float(d.total)} for d in daily_data]

    return render_template('reports/sales.html',
                           sales=sales_data,
                           total_sales=total_sales,
                           total_collected=total_collected,
                           total_credit=total_credit,
                           start_date=start_date,
                           end_date=end_date,
                           payment_type=payment_type,
                           daily_sales=daily_sales)


@app.route('/api/product/<int:id>/details')
@login_required
def product_details(id):
    """جلب تفاصيل المنتج"""
    product = Product.query.get_or_404(id)

    return jsonify({
        'id': product.id,
        'name': product.name,
        'category': product.category,
        'quantity': product.quantity,
        'min_quantity': product.min_quantity,
        'purchase_price': product.purchase_price,
        'selling_price': product.selling_price,
        'location': product.location,
        'is_frozen': product.is_frozen,
        'freeze_deposit': product.freeze_deposit,
        'created_at': product.created_at.strftime('%Y-%m-%d') if product.created_at else None
    })


@app.route('/reports/inventory')
@login_required
@permission_required(Permission.VIEW_REPORTS)
def inventory_report():
    products = Product.query.all()
    total_value = sum(p.quantity * p.purchase_price for p in products)
    low_stock_products = [p for p in products if p.quantity < p.min_quantity]

    return render_template('reports/inventory.html',
                           products=products,
                           total_value=total_value,
                           low_stock_products=low_stock_products)


@app.route('/reports/debts')
@login_required
@permission_required(Permission.VIEW_REPORTS)
def debts_report():
    debt_customers = Customer.query.filter(Customer.balance > 0).all()
    total_debts = sum(c.balance for c in debt_customers)

    return render_template('reports/debts.html',
                           customers=debt_customers,
                           total_debts=total_debts)


# ==================== الحسابات المالية ====================

@app.route('/accounts')
@login_required
@permission_required(Permission.VIEW_REPORTS)
def accounts_list():
    """عرض قائمة الحسابات المالية"""
    accounts = FinancialAccount.query.order_by(FinancialAccount.account_code).all()

    total_assets = sum(a.balance for a in accounts if a.account_type == 'asset')
    total_liabilities = sum(a.balance for a in accounts if a.account_type == 'liability')

    return render_template('accounts/index.html',
                           accounts=accounts,
                           total_assets=total_assets,
                           total_liabilities=total_liabilities)


@app.route('/accounts/trial-balance')
@login_required
@permission_required(Permission.VIEW_REPORTS)
def trial_balance():
    """ميزان المراجعة"""
    accounts = FinancialAccount.query.filter(
        FinancialAccount.account_type.in_(['asset', 'liability', 'equity', 'revenue', 'expense'])
    ).order_by(FinancialAccount.account_code).all()

    total_debit = sum(a.balance for a in accounts if a.account_type in ['asset', 'expense'])
    total_credit = sum(a.balance for a in accounts if a.account_type in ['liability', 'equity', 'revenue'])

    return render_template('accounts/trial_balance.html',
                           accounts=accounts,
                           total_debit=total_debit,
                           total_credit=total_credit)

@app.route('/account_statement', methods=['GET'])
@login_required
@permission_required(Permission.VIEW_REPORTS)
def account_statement():
    account_type = request.args.get('type', 'customer')
    account_id = request.args.get('account_id', type=int)
    from_date = request.args.get('from_date', '')
    to_date = request.args.get('to_date', '')

    customers = Customer.query.order_by(Customer.name.asc()).all()
    suppliers = Supplier.query.order_by(Supplier.name.asc()).all()

    selected_entity = None
    selected_entity_kind = ''
    selected_account_code = ''
    transactions = []
    opening_balance = 0.0
    closing_balance = 0.0
    period_debit = 0.0
    period_credit = 0.0
    entity_current_balance = 0.0

    from_dt = datetime.strptime(from_date, '%Y-%m-%d') if from_date else None
    to_dt = datetime.strptime(to_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59) if to_date else None

    if account_type == 'customer' and account_id:
        selected_entity = db.session.get(Customer, account_id)
        selected_entity_kind = 'العميل'
        if selected_entity:
            selected_account_code = get_customer_posting_account_code(selected_entity.id)

    elif account_type == 'supplier' and account_id:
        selected_entity = db.session.get(Supplier, account_id)
        selected_entity_kind = 'المورد'
        if selected_entity:
            selected_account_code = get_supplier_posting_account_code(selected_entity.id)

    if selected_entity and selected_account_code:
        account_info = get_account_info_by_code(selected_account_code)
        if account_info:
            financial_account_id = account_info[0]
            financial_account = db.session.get(FinancialAccount, financial_account_id)
            entity_current_balance = float(financial_account.balance or 0) if financial_account else 0.0

            # ---------- جلب حركات الفترة ----------
            query = AccountTransaction.query.filter(
                AccountTransaction.account_id == financial_account_id
            )
            if from_dt:
                query = query.filter(AccountTransaction.transaction_date >= from_dt)
            if to_dt:
                query = query.filter(AccountTransaction.transaction_date <= to_dt)

            transactions = query.order_by(
                AccountTransaction.transaction_date.asc(),
                AccountTransaction.id.asc()
            ).all()

            # ---------- حساب الرصيد الافتتاحي (قبل from_date) ----------
            # إذا لم يتم تحديد from_date، الرصيد الافتتاحي = 0
            if from_dt:
                if account_type == 'customer':
                    signed_amount = case(
                        (AccountTransaction.transaction_type == 'debit', AccountTransaction.amount),
                        else_=-AccountTransaction.amount
                    )
                else:  # supplier
                    signed_amount = case(
                        (AccountTransaction.transaction_type == 'credit', AccountTransaction.amount),
                        else_=-AccountTransaction.amount
                    )
                opening_query = db.session.query(
                    func.coalesce(func.sum(signed_amount), 0)
                ).filter(
                    AccountTransaction.account_id == financial_account_id,
                    AccountTransaction.transaction_date < from_dt
                )
                opening_balance = float(opening_query.scalar() or 0)
            else:
                opening_balance = 0.0

            # ---------- حساب الرصيد الجاري والمجاميع ----------
            running = opening_balance
            period_debit = 0.0
            period_credit = 0.0

            for tx in transactions:
                tx.amount_debit = tx.amount if tx.transaction_type == 'debit' else 0
                tx.amount_credit = tx.amount if tx.transaction_type == 'credit' else 0

                # مجاميع العرض (كما هي حسب نوع الحركة)
                if tx.transaction_type == 'debit':
                    period_debit += tx.amount
                else:
                    period_credit += tx.amount

                # تحديث الرصيد الجاري حسب نوع الحساب
                if account_type == 'customer':
                    if tx.transaction_type == 'debit':
                        running += tx.amount
                    else:
                        running -= tx.amount
                else:  # supplier
                    if tx.transaction_type == 'credit':
                        running += tx.amount
                    else:
                        running -= tx.amount

                tx.running_balance = running

            closing_balance = running if transactions else opening_balance

    return render_template('accounts/account_statement.html',
                           account_type=account_type,
                           customers=customers,
                           suppliers=suppliers,
                           selected_entity=selected_entity,
                           selected_entity_kind=selected_entity_kind,
                           selected_account_code=selected_account_code,
                           transactions=transactions,
                           opening_balance=opening_balance,
                           closing_balance=closing_balance,
                           period_debit=period_debit,
                           period_credit=period_credit,
                           entity_current_balance=entity_current_balance,
                           from_date=from_date,
                           to_date=to_date)


def build_account_statement_data(account_type, account_id, from_date='', to_date=''):
    """
    بناء بيانات كشف الحساب باستخدام AccountTransaction (نفس منطق account_statement)
    لضمان تطابق الأرقام مع واجهة المستخدم والتصدير.
    """
    from sqlalchemy import func, case
    from datetime import datetime, time

    customers = Customer.query.order_by(Customer.name.asc()).all()
    suppliers = Supplier.query.order_by(Supplier.name.asc()).all()

    selected_entity = None
    selected_entity_kind = ''
    selected_account_code = ''
    transactions = []
    opening_balance = 0.0
    closing_balance = 0.0
    period_debit = 0.0
    period_credit = 0.0
    current_balance = 0.0

    from_dt = datetime.strptime(from_date, '%Y-%m-%d') if from_date else None
    to_dt = datetime.strptime(to_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59) if to_date else None

    if account_type == 'customer' and account_id:
        selected_entity = db.session.get(Customer, account_id)
        selected_entity_kind = 'العميل'
        if selected_entity:
            selected_account_code = get_customer_posting_account_code(selected_entity.id)
            # الرصيد الحالي من جدول العملاء (أو من الحساب المالي)
            current_balance = float(selected_entity.balance or 0)

    elif account_type == 'supplier' and account_id:
        selected_entity = db.session.get(Supplier, account_id)
        selected_entity_kind = 'المورد'
        if selected_entity:
            selected_account_code = get_supplier_posting_account_code(selected_entity.id)
            current_balance = float(selected_entity.balance or 0)

    if selected_entity and selected_account_code:
        account_info = get_account_info_by_code(selected_account_code)
        if account_info:
            financial_account_id = account_info[0]
            financial_account = db.session.get(FinancialAccount, financial_account_id)

            # ---------- حساب الرصيد الافتتاحي (قبل from_date) ----------
            if from_dt:
                if account_type == 'customer':
                    signed_amount = case(
                        (AccountTransaction.transaction_type == 'debit', AccountTransaction.amount),
                        else_=-AccountTransaction.amount
                    )
                else:  # supplier
                    signed_amount = case(
                        (AccountTransaction.transaction_type == 'credit', AccountTransaction.amount),
                        else_=-AccountTransaction.amount
                    )
                opening_query = db.session.query(
                    func.coalesce(func.sum(signed_amount), 0)
                ).filter(
                    AccountTransaction.account_id == financial_account_id,
                    AccountTransaction.transaction_date < from_dt
                )
                opening_balance = float(opening_query.scalar() or 0)
            else:
                opening_balance = 0.0

            # ---------- جلب حركات الفترة ----------
            query = AccountTransaction.query.filter(
                AccountTransaction.account_id == financial_account_id
            )
            if from_dt:
                query = query.filter(AccountTransaction.transaction_date >= from_dt)
            if to_dt:
                query = query.filter(AccountTransaction.transaction_date <= to_dt)

            account_transactions = query.order_by(
                AccountTransaction.transaction_date.asc(),
                AccountTransaction.id.asc()
            ).all()

            # ---------- تحويل إلى قاموس للتوافق مع القالب ----------
            running = opening_balance
            period_debit = 0.0
            period_credit = 0.0
            for tx in account_transactions:
                amount = float(tx.amount or 0)
                # مجاميع العرض
                if tx.transaction_type == 'debit':
                    period_debit += amount
                else:
                    period_credit += amount

                # تحديث الرصيد الجاري حسب نوع الحساب
                if account_type == 'customer':
                    if tx.transaction_type == 'debit':
                        running += amount
                    else:
                        running -= amount
                else:  # supplier
                    if tx.transaction_type == 'credit':
                        running += amount
                    else:
                        running -= amount

                # إنشاء سجل للقالب (بنفس هيكل القاموس المستخدم سابقاً)
                transactions.append({
                    'date': tx.transaction_date,
                    'reference': f"{tx.reference_type or ''}#{tx.reference_id or ''}",
                    'description': tx.description or '',
                    'debit': tx.amount if tx.transaction_type == 'debit' else 0,
                    'credit': tx.amount if tx.transaction_type == 'credit' else 0,
                    'running_balance': running
                })
            closing_balance = running if account_transactions else opening_balance

    return {
        'customers': customers,
        'suppliers': suppliers,
        'selected_entity': selected_entity,
        'selected_entity_kind': selected_entity_kind,
        'selected_account_code': selected_account_code,
        'transactions': transactions,
        'opening_balance': opening_balance,
        'closing_balance': closing_balance,
        'period_debit': period_debit,
        'period_credit': period_credit,
        'current_balance': current_balance,
        'from_date': from_date,
        'to_date': to_date,
        'account_type': account_type,
        'account_id': account_id,
    }

@app.route('/account_statement/export/excel', methods=['GET'])
@login_required
@permission_required(Permission.VIEW_REPORTS)
def export_account_statement_excel():
    account_type = request.args.get('type', 'customer')
    account_id = request.args.get('account_id', type=int)
    from_date = request.args.get('from_date', '')
    to_date = request.args.get('to_date', '')

    context = build_account_statement_data(account_type, account_id, from_date, to_date)
    entity = context['selected_entity']

    wb = Workbook()
    ws = wb.active
    ws.title = "كشف الحساب"

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="0F172A")
    sub_fill = PatternFill("solid", fgColor="E2E8F0")
    green_fill = PatternFill("solid", fgColor="DCFCE7")
    red_fill = PatternFill("solid", fgColor="FEE2E2")

    title = "كشف حساب العملاء والموردين"
    ws.merge_cells("A1:G1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = "ثلاجة الصليف"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 4
    info_rows = [
        ("نوع الحساب", "عميل" if account_type == "customer" else "مورد"),
        ("الاسم", entity.name if entity else "-"),
        ("الحساب المحاسبي", context["selected_account_code"] or "-"),
        ("الرصيد الحالي", f'{context["current_balance"]:,.2f} ر.ي'),
        ("الرصيد الافتتاحي", f'{context["opening_balance"]:,.2f} ر.ي'),
        ("إجمالي المدين", f'{context["period_debit"]:,.2f} ر.ي'),
        ("إجمالي الدائن", f'{context["period_credit"]:,.2f} ر.ي'),
        ("الرصيد الختامي", f'{context["closing_balance"]:,.2f} ر.ي'),
        ("من تاريخ", from_date or "-"),
        ("إلى تاريخ", to_date or "-"),
    ]

    for label, value in info_rows:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = sub_fill
        ws[f"A{row}"].border = border
        ws[f"B{row}"].border = border
        row += 1

    row += 1
    headers = ["#", "التاريخ", "رقم المرجع", "البيان", "مدين", "دائن", "الرصيد المتحرك"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    row += 1
    for idx, tx in enumerate(context["transactions"], start=1):
        values = [
            idx,
            tx["date"].strftime("%Y-%m-%d %H:%M") if tx.get("date") else "",
            tx.get("reference", "-"),
            tx.get("description", "-"),
            tx.get("debit", 0) or 0,
            tx.get("credit", 0) or 0,
            tx.get("running_balance", 0) or 0,
        ]
        for col, value in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=value)
            c.border = border
            if col in (5, 6, 7):
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal="right")
        if (tx.get("debit") or 0) > 0:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = green_fill if col == 5 else ws.cell(row=row, column=col).fill
        elif (tx.get("credit") or 0) > 0:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = red_fill if col == 6 else ws.cell(row=row, column=col).fill
        row += 1

    for col_letter, width in {
        "A": 6, "B": 20, "C": 18, "D": 48, "E": 14, "F": 14, "G": 16
    }.items():
        ws.column_dimensions[col_letter].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"كشف_حساب_{entity.name if entity else 'غير_محدد'}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/account_statement/export/pdf', methods=['GET'])
@login_required
@permission_required(Permission.VIEW_REPORTS)
def export_account_statement_pdf():
    account_type = request.args.get('type', 'customer')
    account_id = request.args.get('account_id', type=int)
    from_date = request.args.get('from_date', '')
    to_date = request.args.get('to_date', '')

    context = build_account_statement_data(account_type, account_id, from_date, to_date)

    try:
        from weasyprint import HTML
        html = render_template('reports/account_statement_pdf.html', **context)
        pdf = HTML(string=html, base_url=request.host_url).write_pdf()
        filename = f"كشف_حساب_{context['selected_entity'].name if context['selected_entity'] else 'غير_محدد'}.pdf"
        return send_file(
            BytesIO(pdf),
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
    except Exception as e:
        flash(f'تعذر إنشاء PDF: {str(e)}', 'danger')
        return redirect(url_for('account_statement', type=account_type, account_id=account_id, from_date=from_date, to_date=to_date))

# ==================== تشغيل التطبيق ====================
if __name__ == '__main__':
    with app.app_context():
        db.create_all()

        if not User.query.filter_by(username='admin').first():
            admin = User(
                username='admin',
                password=generate_password_hash('admin123'),
                role=UserRole.ADMIN.value,
                full_name='مدير النظام',
                phone='123456789',
                email='admin@thaljat.com'
            )
            db.session.add(admin)
            db.session.flush()

            admin_employee = Employee(
                full_name='مدير النظام',
                position='مدير',
                phone='123456789',
                email='admin@thaljat.com',
                salary=5000,
                user_id=admin.id
            )
            db.session.add(admin_employee)

            db.session.commit()
            print("=" * 50)
            print("تم إنشاء المستخدم admin بنجاح")
            print("اسم المستخدم: admin")
            print("كلمة المرور: admin123")
            print("=" * 50)

    # للتشغيل المحلي
    # app.run(debug=True)

    # للنشر على Render
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))